import csv
import io
import os
import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from main import run_pipeline
from modules.url_source import (
    get_urls_from_discovery,
    VALID_STATES,
    DISCOVERY_TOPICS,
    _load_existing_domains_from_excel,
    _extract_domain,
    _build_merged_workbook,
)
from config import DEFAULT_TEMPERATURE, DEFAULT_TRUNCATION, ERRORS_CSV, MARKDOWN_CSV

# ── Hardcoded pipeline defaults ───────────────────────────────────────────────
DEFAULT_PROVIDER       = "ollama"
DEFAULT_MODEL          = "qwen2.5:7b"
DEFAULT_TEMP           = DEFAULT_TEMPERATURE
DEFAULT_TRUNCATION_VAL = DEFAULT_TRUNCATION

# Accepted URL column names — case-insensitive
URL_COLUMN_NAMES = {"url", "urls", "links", "link", "website", "websites"}

st.set_page_config(page_title="IncentivAI", layout="wide")
st.title("IncentivAI – Utility Incentive Extractor")

# ── Session state init ────────────────────────────────────────────────────────
for key, default in {
    "cancelled":        False,
    "output_file":      None,
    "markdown_done":    False,
    "errors_done":      False,
    "run_complete":     False,
    "discovery_result": None,
    "merge_result":     None,
    "single_done":      False,
    "md_input_done":    False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("Configuration")

mode = st.sidebar.radio(
    "Select Mode",
    ["Upload Excel", "Single URL", "Upload Markdown", "City URL Discovery"]
)

# ── Sidebar inputs per mode ───────────────────────────────────────────────────
uploaded_file  = None
single_url_in  = None
uploaded_md    = None

# Safe defaults
temperature       = DEFAULT_TEMP
truncation_length = DEFAULT_TRUNCATION_VAL
provider          = DEFAULT_PROVIDER
model_name        = DEFAULT_MODEL
run_button        = False
cancel_button     = False

if mode == "Upload Excel":
    uploaded_file = st.sidebar.file_uploader(
        "Upload Excel (.xlsx) — column can be named URLs, urls, links, etc.",
        type=["xlsx"]
    )

elif mode == "Single URL":
    single_url_in = st.sidebar.text_input("Enter a single URL to run the pipeline on")

elif mode == "Upload Markdown":
    uploaded_md = st.sidebar.file_uploader(
        "Upload a .md or .txt file to run the pipeline on",
        type=["md", "txt"]
    )

# Pipeline settings shown for all extraction modes
if mode in ("Upload Excel", "Single URL", "Upload Markdown"):
    temperature       = st.sidebar.number_input("Temperature",       value=DEFAULT_TEMP,           step=0.1)
    truncation_length = st.sidebar.number_input("Max Scrape Length", value=DEFAULT_TRUNCATION_VAL, step=1000)
    provider          = st.sidebar.selectbox(
        "LLM Provider",
        ["ollama", "openai", "uw_ssec", "anthropic", "google"],
        index=0
    )
    model_name    = st.sidebar.text_input("Model Name", value=DEFAULT_MODEL)
    run_button    = st.sidebar.button("▶ Run Extraction")
    cancel_button = st.sidebar.button("⏹ Cancel")

    if cancel_button:
        st.session_state.cancelled = True
        st.sidebar.warning("Cancellation requested — stopping after current URL.")


# ── Helper: detect URL column in Excel ───────────────────────────────────────
def _find_url_column(columns):
    """Find the URL column regardless of capitalisation."""
    for col in columns:
        if str(col).strip().lower() in URL_COLUMN_NAMES:
            return col
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: City URL Discovery
# ═══════════════════════════════════════════════════════════════════════════════
if mode == "City URL Discovery":
    st.subheader("Discover New Utility URLs by State")
    st.caption(
        "Searches OpenSERP for electric utility and cooperative websites by state. "
        "Skips domains already in your existing database."
    )

    disc_col1, disc_col2 = st.columns(2)
    with disc_col1:
        selected_states = st.multiselect("States to search", VALID_STATES, default=["Texas"])
        openserp_url    = st.text_input("OpenSERP URL", value="http://localhost:7070")
        engine          = st.selectbox("Search engine", ["google", "bing", "duckduckgo"], index=0)
    with disc_col2:
        num_results = st.slider("Results per query", min_value=3, max_value=15, value=8)
        db_file     = st.file_uploader(
            "Existing URL database for deduplication (optional)",
            type=["xlsx"],
            key="db_upload"
        )

    with st.expander("Search topics"):
        topics_text = st.text_area(
            "One topic per line",
            value="\n".join(DISCOVERY_TOPICS),
            height=300
        )

    if st.button("▶ Run Discovery", disabled=not selected_states):
        progress_bar = st.progress(0)
        status_text  = st.empty()

        def discovery_progress(current, total, url="", message=""):
            progress_bar.progress(current / total)
            status_text.markdown(f"**{message}**")

        discovered = get_urls_from_discovery(
            states=selected_states,
            openserp_url=openserp_url,
            engine=engine,
            num_results=num_results,
            existing_db=db_file,
            progress_callback=discovery_progress,
        )
        st.session_state.discovery_result = discovered

    # ── Show discovery results (persists across reruns) ───────────────────────
    discovered = st.session_state.discovery_result
    if discovered is not None:
        st.success(f"Found **{len(discovered)}** new utility URLs.")

        if discovered:
            st.subheader("Discovered URLs")
            df = pd.DataFrame([{
                "State":         r["state"],
                "URL":           r["url"],
                "Page Title":    r["title"],
                "Discovered At": r["discovered_at"],
            } for r in discovered])
            st.dataframe(df, use_container_width=True)

            def _thin():
                s = Side(style="thin", color="BFBFBF")
                return Border(left=s, right=s, top=s, bottom=s)

            wb = Workbook()
            ws = wb.active
            ws.title = "Discovered URLs"
            for col, w in {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}.items():
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"
            for c, h in enumerate(["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"], 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin()

            for r_idx, row in enumerate(discovered, 2):
                for c, val in enumerate([
                    row["state"], row["query"], row["url"],
                    row["title"], row["description"], row["discovered_at"]
                ], 1):
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = _thin()
                    cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
                    if c == 1:
                        cell.fill = PatternFill("solid", start_color="D6E4F0")
                        cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                "⬇️ Download Discovered URLs Excel",
                buf.read(),
                file_name=f"utility_urls_discovered_{datetime.date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            url_list = "\n".join(r["url"] for r in discovered)
            st.download_button(
                "⬇️ Download as plain URL list (.txt)",
                url_list.encode("utf-8"),
                file_name="discovered_urls.txt",
                mime="text/plain"
            )

    # ── Merge sub-section ─────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Merge Discovered URLs into Existing Database")
    st.caption("Domain-level deduplication. Merges a discovered file into your existing URL database.")

    merge_col1, merge_col2 = st.columns(2)
    with merge_col1:
        merge_db   = st.file_uploader("Existing URL database",  type=["xlsx"], key="merge_db")
    with merge_col2:
        merge_disc = st.file_uploader("Discovered URLs file",   type=["xlsx"], key="merge_disc")

    if merge_db and merge_disc:
        from openpyxl import load_workbook as lw

        def _load_urls(f) -> list:
            wb = lw(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            urls = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and cell.startswith("http"):
                        urls.append(cell.strip())
            wb.close()
            return urls

        def _load_disc_rows(f) -> list:
            wb = lw(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            rows    = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(headers, row))
                url = str(record.get("URL", "") or "").strip()
                if url.startswith("http"):
                    rows.append(record)
            wb.close()
            return rows

        existing_urls    = _load_urls(merge_db)
        existing_domains = set(_extract_domain(u) for u in existing_urls if _extract_domain(u))
        discovered_rows  = _load_disc_rows(merge_disc)

        seen     = set(existing_domains)
        new_rows = []
        skipped  = 0
        for row in discovered_rows:
            url    = str(row.get("URL", "")).strip()
            domain = _extract_domain(url)
            if not domain:
                continue
            if domain in seen:
                skipped += 1
            else:
                seen.add(domain)
                new_rows.append(row)

        st.write(f"**Existing:** {len(existing_urls)} URLs | **Discovered:** {len(discovered_rows)} rows")
        st.write(f"**After dedup:** {len(new_rows)} new URLs, {skipped} skipped")

        if new_rows and st.button("▶ Merge & Download"):
            wb  = _build_merged_workbook(existing_urls, new_rows)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success(f"Merged: {len(existing_urls)} existing + {len(new_rows)} new = {len(existing_urls)+len(new_rows)} total")
            st.download_button(
                "⬇️ Download Merged Database",
                buf.read(),
                file_name=f"Relevant_URLs_merged_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: Upload Excel / Single URL / Upload Markdown — Extraction pipeline
# ═══════════════════════════════════════════════════════════════════════════════
elif mode in ("Upload Excel", "Single URL", "Upload Markdown"):

    tab_progress, tab_markdown, tab_errors = st.tabs([
        "📊 Progress", "📝 Live Summaries", "⚠️ Errors"
    ])

    if run_button:
        st.session_state.cancelled    = False
        st.session_state.run_complete = False
        st.session_state.output_file  = None

        # ── Validate inputs ───────────────────────────────────────────────────
        error = None
        if mode == "Upload Excel" and not uploaded_file:
            error = "Please upload an Excel file."
        elif mode == "Single URL" and not single_url_in:
            error = "Please enter a URL."
        elif mode == "Upload Markdown" and not uploaded_md:
            error = "Please upload a markdown or text file."

        if error:
            st.error(error)
        else:
            # ── Resolve inputs into what run_pipeline expects ─────────────────
            pipeline_mode    = "Upload Excel"
            pipeline_file    = None
            pipeline_state   = None
            markdown_content = None

            if mode == "Upload Excel":
                df_preview = pd.read_excel(uploaded_file)
                url_col    = _find_url_column(df_preview.columns)
                if url_col is None:
                    st.error(
                        f"No URL column found. Expected one of: {', '.join(sorted(URL_COLUMN_NAMES))}. "
                        f"Found: {', '.join(df_preview.columns.tolist())}"
                    )
                    st.stop()
                if url_col != "URLs":
                    df_preview = df_preview.rename(columns={url_col: "URLs"})
                    buf = io.BytesIO()
                    df_preview.to_excel(buf, index=False)
                    buf.seek(0)
                    pipeline_file = buf
                else:
                    pipeline_file = uploaded_file
                pipeline_mode = "Upload Excel"

            elif mode == "Single URL":
                df_single = pd.DataFrame({"URLs": [single_url_in.strip()]})
                buf = io.BytesIO()
                df_single.to_excel(buf, index=False)
                buf.seek(0)
                pipeline_file = buf
                pipeline_mode = "Upload Excel"

            elif mode == "Upload Markdown":
                markdown_content = uploaded_md.read().decode("utf-8", errors="replace")

            with tab_progress:
                progress_bar = st.progress(0)
                status_text  = st.empty()
                stats        = st.empty()

            success_count = [0]
            fail_count    = [0]

            def progress_callback(current, total, url="", message=""):
                pct = current / total if total > 0 else 0
                progress_bar.progress(pct)
                status_text.markdown(f"**{message}**")
                stats.markdown(
                    f"✅ Succeeded: `{success_count[0]}` &nbsp;|&nbsp; "
                    f"❌ Failed: `{fail_count[0]}` &nbsp;|&nbsp; "
                    f"🔗 Current: `{url}`"
                )
                with tab_markdown:
                    if os.path.isfile(MARKDOWN_CSV):
                        try:
                            md_df = pd.read_csv(MARKDOWN_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                            for _, row in md_df.tail(5).iterrows():
                                st.markdown(row["markdown_summary"])
                                st.markdown("---")
                        except Exception:
                            pass
                with tab_errors:
                    if os.path.isfile(ERRORS_CSV):
                        try:
                            err_df = pd.read_csv(ERRORS_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                            st.dataframe(err_df, use_container_width=True)
                        except Exception:
                            pass

            # ── Run pipeline ──────────────────────────────────────────────────
            if mode == "Upload Markdown":
                from modules.processor import process_text
                from modules.exporter import export_to_csv, append_markdown_entry
                import datetime as dt

                with tab_progress:
                    status_text.markdown("**Running LLM extraction on uploaded markdown...**")

                try:
                    structured = process_text(markdown_content, "uploaded_markdown", temperature)
                    structured["source_url"]           = uploaded_md.name
                    structured["parent_url"]           = None
                    structured["is_sublink"]           = False
                    structured["url_type"]             = "markdown"
                    structured["extraction_timestamp"] = dt.datetime.utcnow().isoformat()
                    append_markdown_entry(structured, MARKDOWN_CSV)
                    output_file = export_to_csv([structured])
                    st.session_state.output_file  = output_file
                    st.session_state.run_complete = True
                    st.session_state.md_input_done = True
                    st.session_state.md_input_text = markdown_content
                    st.session_state.md_input_name = uploaded_md.name

                except Exception as e:
                    st.error(f"Extraction failed: {e}")

            else:
                output_file = run_pipeline(
                    mode=pipeline_mode,
                    uploaded_file=pipeline_file,
                    state=pipeline_state,
                    temperature=temperature,
                    truncation_length=int(truncation_length),
                    progress_callback=progress_callback,
                    cancel_flag=lambda: st.session_state.cancelled,
                    provider=provider,
                    model=model_name,
                )
                st.session_state.output_file  = output_file
                st.session_state.run_complete = True

    # ── Post-run results (persists via session state) ─────────────────────────
    if st.session_state.run_complete and st.session_state.output_file:
        output_file = st.session_state.output_file

        with tab_progress:
            if st.session_state.cancelled:
                st.warning("Pipeline was cancelled. Partial results saved.")
            else:
                st.success("✅ Extraction complete.")

            with open(output_file, "rb") as f:
                st.download_button(
                    "⬇️ Download Results CSV",
                    f,
                    file_name="incentives_output.csv"
                )

            if st.session_state.get("md_input_done"):
                st.download_button(
                    "⬇️ Download Original Markdown",
                    st.session_state.md_input_text.encode("utf-8"),
                    file_name=st.session_state.md_input_name,
                    mime="text/markdown"
                )

        with tab_markdown:
            st.subheader("Full Markdown Summaries")
            if os.path.isfile(MARKDOWN_CSV):
                try:
                    md_df = pd.read_csv(MARKDOWN_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")

                    all_md = "\n\n---\n\n".join(md_df["markdown_summary"].dropna().tolist())

                    st.download_button(
                        "⬇️ Download All Summaries as Markdown",
                        all_md.encode("utf-8"),
                        file_name="incentive_summaries.md",
                        mime="text/markdown"
                    )
                    with open(MARKDOWN_CSV, "rb") as f:
                        st.download_button(
                            "⬇️ Download Markdown CSV",
                            f,
                            file_name="markdown_summaries.csv"
                        )

                    for _, row in md_df.iterrows():
                        st.markdown(row["markdown_summary"])
                        st.markdown("---")

                except Exception as e:
                    st.error(f"Could not load markdown summaries: {e}")
            else:
                st.info("No summaries generated yet.")

        with tab_errors:
            st.subheader("Error Log")
            if os.path.isfile(ERRORS_CSV):
                try:
                    err_df = pd.read_csv(ERRORS_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                    st.dataframe(err_df, use_container_width=True)
                    with open(ERRORS_CSV, "rb") as f:
                        st.download_button(
                            "⬇️ Download Error Log",
                            f,
                            file_name="errors.csv"
                        )
                except Exception as e:
                    st.error(f"Could not load error log: {e}")
            else:
                st.info("No errors logged.")
