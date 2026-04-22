"""
Discover electric utility incentive URLs by U.S. state via local OpenSERP instance.
Goal: find NEW utility/cooperative/municipal-energy official websites offering
      rebates, grants, or incentive programs — to be merged into the DSIRE URL database.

Results are saved to an Excel file (one row per search result URL).

Usage:
    python energy_search.py                        # interactive
    python energy_search.py Texas                  # single state
    python energy_search.py "New York" California  # multiple states, appends to same file
"""

import re
import sys
import time
import datetime
import requests
from typing import List, Dict, Any
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not found — run: pip install openpyxl")

# ── Configuration ─────────────────────────────────────────────────────────────
OPENSERP_URL  = "http://localhost:7070"
SEARCH_ENGINE = "google"
NUM_RESULTS   = 8       # more results per query to maximize new URL discovery
COUNTRY       = "us"
LANGUAGE      = "EN"
OUTPUT_FILE   = "utility_urls_discovered.xlsx"

# ── Search topics ─────────────────────────────────────────────────────────────
# Goal: surface UTILITY OFFICIAL WEBSITES (not news/aggregators).
# Each phrase + state name => Google query.
# Designed to match domain types seen in DSIRE:
#   .coop, municipalelectric.*, *electric.com, energy.*.gov, pud.*.*, etc.
TOPICS = [
    # Find utility/coop official sites directly
    "electric cooperative official website rebate",
    "municipal electric utility rebate program",
    "public utility district energy rebate",
    "rural electric cooperative incentive program",
    "investor owned utility energy efficiency rebate",

    # Device-specific — utilities publish these as standalone program pages
    "utility solar panel rebate apply",
    "utility heat pump rebate program apply",
    "utility EV charger rebate apply",
    "utility smart thermostat rebate program",
    "utility battery storage incentive apply",
    "utility weatherization rebate low income",
    "utility net metering program apply",
    "utility on-bill financing energy upgrade",
    "utility demand response incentive program",
    "utility energy efficiency rebate commercial",
]

# ── Domain filter ─────────────────────────────────────────────────────────────
# Blocklist: aggregators, news, advocacy, govt-non-utility, short links
# A result is KEPT if it does NOT match any of these patterns.
DOMAIN_BLOCKLIST = re.compile(
    r"(dsire|energysage|energystar|epa\.gov|energy\.gov$"
    r"|nrel\.gov|eia\.gov|wikipedia|energycoalition"
    r"|nrdc\.org|sierraclub|greentechmedia|pv-magazine"
    r"|forbes|bloomberg|reuters|apnews|cnn\.com|nytimes"
    r"|bit\.ly|tinyurl|t\.co"
    r"|bcap-ocean|cleanairfleets|bcapcodes"
    r"|mgaleg|comptroller|sos\."          # legislative / secretary of state
    r")",
    re.IGNORECASE,
)

def is_utility_url(url: str) -> bool:
    """
    Heuristic: keep URLs that look like utility/government-energy official sites.
    Drops aggregators, news sites, and known non-utility domains.
    """
    try:
        domain = urlparse(url).netloc.lower()
    except Exception:
        return False
    if DOMAIN_BLOCKLIST.search(domain):
        return False
    return True

# ── Excel helpers ─────────────────────────────────────────────────────────────
COL_WIDTHS = {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _setup_sheet(ws):
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

def _write_header(ws):
    headers = ["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"]
    border = _thin_border()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 20

def _append_row(ws, state, query, url, title, desc, discovered_at):
    border = _thin_border()
    row_data = [state, query, url, title, desc, discovered_at]
    r = ws.max_row + 1
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border    = border
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
        if c == 1:   # State — highlight
            cell.fill = PatternFill("solid", start_color="D6E4F0")
            cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    ws.row_dimensions[r].height = 45

def _get_or_create_workbook(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws  = wb.active
        ws.title = "Discovered URLs"
        _setup_sheet(ws)
        _write_header(ws)
    return wb, ws

# ── Core search ───────────────────────────────────────────────────────────────
def search(query: str, limit: int = NUM_RESULTS) -> List[Dict[str, Any]]:
    url    = f"{OPENSERP_URL}/{SEARCH_ENGINE}/search"
    params = {"text": query, "limit": limit, "gl": COUNTRY, "lang": LANGUAGE}
    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  ❌ Search error: {e}")
        return []

# ── Per-state search ──────────────────────────────────────────────────────────
def run_state(state: str) -> None:
    print(f"\n🔍 Discovering utility URLs for: {state}")
    wb, ws = (_get_or_create_workbook(OUTPUT_FILE) if EXCEL_AVAILABLE else (None, None))

    total_found = 0
    total_kept  = 0

    for topic in TOPICS:
        query       = f"{topic} {state}"
        discovered  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\n  📌 {query}")

        results = search(query)
        if not results:
            print("     No results.")
            time.sleep(1)
            continue

        kept = []
        for item in results:
            url = item.get("url", "")
            if is_utility_url(url):
                kept.append(item)
            else:
                print(f"     ⛔ filtered: {url}")

        total_found += len(results)
        total_kept  += len(kept)

        if not kept:
            print("     No utility URLs after filtering.")
        else:
            for item in kept:
                url   = item.get("url", "")
                title = item.get("title", "")
                desc  = item.get("description", "")
                print(f"     ✅ {title[:60]}")
                print(f"        {url}")
                if EXCEL_AVAILABLE and ws is not None:
                    _append_row(ws, state, query, url, title, desc, discovered)

        time.sleep(1)

    if EXCEL_AVAILABLE and wb is not None:
        wb.save(OUTPUT_FILE)

    print(f"\n  📊 {state}: {total_kept}/{total_found} URLs kept → {OUTPUT_FILE}")

# ── Entry points ──────────────────────────────────────────────────────────────
VALID_STATES = {
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
}

def _interactive():
    print(f"🔍 Utility URL Discovery\nOpenSERP: {OPENSERP_URL} | Engine: {SEARCH_ENGINE}\n")
    while True:
        state = input("🗺️  Enter a U.S. state: ").strip().title()
        if state in VALID_STATES:
            break
        print(f"  ⚠️  '{state}' not recognized. Try again.")
    run_state(state)

def _cli(states: list):
    for raw in states:
        state = raw.strip().title()
        if state not in VALID_STATES:
            print(f"⚠️  Skipping '{state}' — not a recognized U.S. state.")
            continue
        print(f"\n{'='*52}\n  🗺️  {state}\n{'='*52}")
        run_state(state)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        _cli(sys.argv[1:])
    else:
        _interactive()
