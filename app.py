
# nest_asyncio MUST be the very first import and applied before anything else.
import nest_asyncio
nest_asyncio.apply()

import csv
import io
import os
import datetime
import warnings
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Suppress crawl4ai background task noise
warnings.filterwarnings("ignore", message="Task was destroyed but it is pending")

from main import run_pipeline
from modules.url_source import (
    get_urls_from_discovery,
    VALID_STATES,
    DISCOVERY_TOPICS,
    _load_existing_domains_from_excel,
    _extract_domain,
    _build_merged_workbook,
)
from config import DEFAULT_TEMPERATURE, DEFAULT_TRUNCATION, ERRORS_CSV, MARKDOWN_CSV

# ── Hardcoded pipeline defaults ───────────────────────────────────────────────
DEFAULT_PROVIDER       = "ollama"
DEFAULT_MODEL          = "qwen2.5:14b"
DEFAULT_TEMP           = DEFAULT_TEMPERATURE
DEFAULT_TRUNCATION_VAL = DEFAULT_TRUNCATION
DEFAULT_MAX_DEPTH      = 3

URL_COLUMN_NAMES = {"url", "urls", "links", "link", "website", "websites"}

st.set_page_config(page_title="IncentivAI", layout="wide")
st.title("IncentivAI – Utility Incentive Extractor")

# ── Session state init ────────────────────────────────────────────────────────
for key, default in {
    "cancelled":        False,
    "output_file":      None,
    "markdown_done":    False,
    "errors_done":      False,
    "run_complete":     False,
    "discovery_result": None,
    "merge_result":     None,
    "single_done":      False,
    "md_input_done":    False,
    "callback_count":   0,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ── Version-safe scrollable container ────────────────────────────────────────
def _scrollable_container(height: int = 600):
    try:
        return st.container(height=height, border=True)
    except TypeError:
        return st.expander("📝 Summaries", expanded=True)


# ── Version-safe st.dataframe ─────────────────────────────────────────────────
def _dataframe(df: pd.DataFrame):
    df = _safe_dataframe(df)
    try:
        st.dataframe(df, width="stretch")
    except TypeError:
        st.dataframe(df, use_container_width=True)


# ── PyArrow LargeUtf8 fix ─────────────────────────────────────────────────────
def _safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            try:
                df[col] = df[col].astype("string").fillna("")
            except Exception:
                df[col] = df[col].astype(str).astype("string").fillna("")
    return df


def _read_csv_safe(path: str):
    if not os.path.isfile(path):
        return None
    try:
        df = pd.read_csv(path, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
        return _safe_dataframe(df)
    except Exception:
        return None


# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("Configuration")

mode = st.sidebar.radio(
    "Select Mode",
    ["Upload Excel", "Single URL", "Upload Markdown", "City URL Discovery"]
)

uploaded_file  = None
single_url_in  = None
uploaded_md    = None

temperature       = DEFAULT_TEMP
truncation_length = DEFAULT_TRUNCATION_VAL
provider          = DEFAULT_PROVIDER
model_name        = DEFAULT_MODEL
max_depth         = DEFAULT_MAX_DEPTH
run_button        = False
cancel_button     = False

if mode == "Upload Excel":
    uploaded_file = st.sidebar.file_uploader(
        "Upload Excel (.xlsx) — column can be named URLs, urls, links, etc.",
        type=["xlsx"]
    )
elif mode == "Single URL":
    single_url_in = st.sidebar.text_input("Enter a single URL to run the pipeline on")
elif mode == "Upload Markdown":
    uploaded_md = st.sidebar.file_uploader(
        "Upload a .md or .txt file to run the pipeline on",
        type=["md", "txt"]
    )

if mode in ("Upload Excel", "Single URL", "Upload Markdown"):
    temperature       = st.sidebar.number_input("Temperature",       value=DEFAULT_TEMP,           step=0.1)
    truncation_length = st.sidebar.number_input("Max Scrape Length", value=DEFAULT_TRUNCATION_VAL, step=1000)
    provider          = st.sidebar.selectbox(
        "LLM Provider",
        ["ollama", "openai", "uw_ssec", "anthropic", "google"],
        index=0
    )
    model_name = st.sidebar.text_input("Model Name", value=DEFAULT_MODEL)

    # ── Crawl depth slider ────────────────────────────────────────────────────
    max_depth = st.sidebar.slider(
        "Crawl Depth",
        min_value=1,
        max_value=5,
        value=DEFAULT_MAX_DEPTH,
        help=(
            "1 = scrape only the exact URL given, no sublinks. "
            "3 = follow links up to 3 levels deep (default). "
            "Higher = more pages found but slower and more memory."
        )
    )

    if max_depth == 1:
        st.sidebar.caption("⚡ Depth 1 — single page only, fastest.")
    elif max_depth <= 2:
        st.sidebar.caption("🔍 Depth 2 — page + immediate sublinks.")
    elif max_depth == 3:
        st.sidebar.caption("🔍 Depth 3 — default, balanced coverage.")
    else:
        st.sidebar.caption("⚠️ Deep crawl — slow, high memory usage.")

    run_button    = st.sidebar.button("▶ Run Extraction")
    cancel_button = st.sidebar.button("⏹ Cancel")

    if cancel_button:
        st.session_state.cancelled = True
        st.sidebar.warning("Cancellation requested — stopping after current URL.")


def _find_url_column(columns):
    for col in columns:
        if str(col).strip().lower() in URL_COLUMN_NAMES:
            return col
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: City URL Discovery
# ═══════════════════════════════════════════════════════════════════════════════
if mode == "City URL Discovery":
    st.subheader("Discover New Utility URLs by State")
    st.caption(
        "Searches OpenSERP for electric utility and cooperative websites by state. "
        "Skips domains already in your existing database."
    )

    disc_col1, disc_col2 = st.columns(2)
    with disc_col1:
        selected_states = st.multiselect("States to search", VALID_STATES, default=["Texas"])
        openserp_url    = st.text_input("OpenSERP URL", value="http://localhost:7070")
        engine          = st.selectbox("Search engine", ["bing", "duckduckgo", "google"], index=0)
    with disc_col2:
        num_results = st.slider("Results per query", min_value=3, max_value=15, value=8)
        db_file     = st.file_uploader(
            "Existing URL database for deduplication (optional)",
            type=["xlsx"],
            key="db_upload"
        )

    with st.expander("Search topics"):
        topics_text = st.text_area(
            "One topic per line",
            value="\n".join(DISCOVERY_TOPICS),
            height=300
        )

    if st.button("▶ Run Discovery", disabled=not selected_states):
        progress_bar = st.progress(0)
        status_text  = st.empty()

        def discovery_progress(current, total, url="", message=""):
            progress_bar.progress(current / total)
            status_text.markdown(f"**{message}**")

        discovered = get_urls_from_discovery(
            states=selected_states,
            openserp_url=openserp_url,
            engine=engine,
            num_results=num_results,
            existing_db=db_file,
            progress_callback=discovery_progress,
        )
        st.session_state.discovery_result = discovered

    discovered = st.session_state.discovery_result
    if discovered is not None:
        st.success(f"Found **{len(discovered)}** new utility URLs.")

        if discovered:
            st.subheader("Discovered URLs")
            df = pd.DataFrame([{
                "State":         r["state"],
                "URL":           r["url"],
                "Page Title":    r["title"],
                "Discovered At": r["discovered_at"],
            } for r in discovered])
            _dataframe(df)

            def _thin():
                s = Side(style="thin", color="BFBFBF")
                return Border(left=s, right=s, top=s, bottom=s)

            wb = Workbook()
            ws = wb.active
            ws.title = "Discovered URLs"
            for col, w in {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}.items():
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"
            for c, h in enumerate(["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"], 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin()

            for r_idx, row in enumerate(discovered, 2):
                for c, val in enumerate([
                    row["state"], row["query"], row["url"],
                    row["title"], row["description"], row["discovered_at"]
                ], 1):
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = _thin()
                    cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
                    if c == 1:
                        cell.fill = PatternFill("solid", start_color="D6E4F0")
                        cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.download_button(
                "⬇️ Download Discovered URLs Excel",
                buf.read(),
                file_name=f"utility_urls_discovered_{datetime.date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            url_list = "\n".join(r["url"] for r in discovered)
            st.download_button(
                "⬇️ Download as plain URL list (.txt)",
                url_list.encode("utf-8"),
                file_name="discovered_urls.txt",
                mime="text/plain"
            )

    st.markdown("---")
    st.subheader("Merge Discovered URLs into Existing Database")
    st.caption("Domain-level deduplication. Merges a discovered file into your existing URL database.")

    merge_col1, merge_col2 = st.columns(2)
    with merge_col1:
        merge_db   = st.file_uploader("Existing URL database",  type=["xlsx"], key="merge_db")
    with merge_col2:
        merge_disc = st.file_uploader("Discovered URLs file",   type=["xlsx"], key="merge_disc")

    if merge_db and merge_disc:
        from openpyxl import load_workbook as lw

        def _load_urls(f) -> list:
            wb = lw(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            urls = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and cell.startswith("http"):
                        urls.append(cell.strip())
            wb.close()
            return urls

        def _load_disc_rows(f) -> list:
            wb = lw(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            rows    = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(headers, row))
                url = str(record.get("URL", "") or "").strip()
                if url.startswith("http"):
                    rows.append(record)
            wb.close()
            return rows

        existing_urls    = _load_urls(merge_db)
        existing_domains = set(_extract_domain(u) for u in existing_urls if _extract_domain(u))
        discovered_rows  = _load_disc_rows(merge_disc)

        seen     = set(existing_domains)
        new_rows = []
        skipped  = 0
        for row in discovered_rows:
            url    = str(row.get("URL", "")).strip()
            domain = _extract_domain(url)
            if not domain:
                continue
            if domain in seen:
                skipped += 1
            else:
                seen.add(domain)
                new_rows.append(row)

        st.write(f"**Existing:** {len(existing_urls)} URLs | **Discovered:** {len(discovered_rows)} rows")
        st.write(f"**After dedup:** {len(new_rows)} new URLs, {skipped} skipped")

        if new_rows and st.button("▶ Merge & Download"):
            wb  = _build_merged_workbook(existing_urls, new_rows)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success(f"Merged: {len(existing_urls)} existing + {len(new_rows)} new = {len(existing_urls)+len(new_rows)} total")
            st.download_button(
                "⬇️ Download Merged Database",
                buf.read(),
                file_name=f"Relevant_URLs_merged_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: Upload Excel / Single URL / Upload Markdown
# ═══════════════════════════════════════════════════════════════════════════════
elif mode in ("Upload Excel", "Single URL", "Upload Markdown"):

    tab_progress, tab_markdown, tab_errors = st.tabs([
        "📊 Progress", "📝 Live Summaries", "⚠️ Errors"
    ])

    if run_button:
        st.session_state.cancelled    = False
        st.session_state.run_complete = False
        st.session_state.output_file  = None
        st.session_state.callback_count = 0

        error = None
        if mode == "Upload Excel" and not uploaded_file:
            error = "Please upload an Excel file."
        elif mode == "Single URL" and not single_url_in:
            error = "Please enter a URL."
        elif mode == "Upload Markdown" and not uploaded_md:
            error = "Please upload a markdown or text file."

        if error:
            st.error(error)
        else:
            pipeline_mode  = "Upload Excel"
            pipeline_file  = None
            pipeline_state = None

            if mode == "Upload Excel":
                df_preview = pd.read_excel(uploaded_file)
                url_col    = _find_url_column(df_preview.columns)
                if url_col is None:
                    st.error(
                        f"No URL column found. Expected one of: {', '.join(sorted(URL_COLUMN_NAMES))}. "
                        f"Found: {', '.join(df_preview.columns.tolist())}"
                    )
                    st.stop()
                if url_col != "URLs":
                    df_preview = df_preview.rename(columns={url_col: "URLs"})
                buf = io.BytesIO()
                df_preview.to_excel(buf, index=False)
                buf.seek(0)
                pipeline_file = buf
                pipeline_mode = "Upload Excel"

            elif mode == "Single URL":
                df_single = pd.DataFrame({"URLs": [single_url_in.strip()]})
                buf = io.BytesIO()
                df_single.to_excel(buf, index=False)
                buf.seek(0)
                pipeline_file = buf
                pipeline_mode = "Upload Excel"

            elif mode == "Upload Markdown":
                markdown_content = uploaded_md.read().decode("utf-8", errors="replace")
                pipeline_mode    = "Upload Markdown"
                pipeline_file    = io.StringIO(markdown_content)
                pipeline_file.name = uploaded_md.name

            with tab_progress:
                progress_bar = st.progress(0)
                status_text  = st.empty()
                stats        = st.empty()
                live_summary = st.empty()

            success_count = [0]
            fail_count    = [0]

            def progress_callback(current, total, url="", message=""):
                st.session_state.callback_count += 1
                pct = current / total if total > 0 else 0
                progress_bar.progress(pct)
                status_text.markdown(f"**{message}**")
                stats.markdown(
                    f"✅ Succeeded: `{success_count[0]}` &nbsp;|&nbsp; "
                    f"❌ Failed: `{fail_count[0]}` &nbsp;|&nbsp; "
                    f"🔗 Current: `{url}`"
                )

                # Throttle UI updates — only refresh every 3 callbacks
                if st.session_state.callback_count % 3 == 0:
                    md_df = _read_csv_safe(MARKDOWN_CSV)
                    if md_df is not None and "markdown_summary" in md_df.columns:
                        last_3 = md_df["markdown_summary"].dropna().tail(3).tolist()
                        with live_summary:
                            for summary in last_3:
                                st.markdown(summary)
                                st.markdown("---")

                    err_df = _read_csv_safe(ERRORS_CSV)
                    if err_df is not None:
                        with tab_errors:
                            _dataframe(err_df)

            output_file = run_pipeline(
                mode=pipeline_mode,
                uploaded_file=pipeline_file,
                state=pipeline_state,
                temperature=temperature,
                truncation_length=int(truncation_length),
                progress_callback=progress_callback,
                cancel_flag=lambda: st.session_state.cancelled,
                provider=provider,
                model=model_name,
                max_depth=int(max_depth),
            )
            st.session_state.output_file  = output_file
            st.session_state.run_complete = True

    # ── Post-run results ──────────────────────────────────────────────────────
    if st.session_state.run_complete and st.session_state.output_file:
        output_file = st.session_state.output_file

        with tab_progress:
            if st.session_state.cancelled:
                st.warning("Pipeline was cancelled. Partial results saved.")
            else:
                st.success("✅ Extraction complete.")

            with open(output_file, "rb") as f:
                st.download_button(
                    "⬇️ Download Results CSV",
                    f,
                    file_name="incentives_output.csv"
                )

        with tab_markdown:
            st.subheader("Full Markdown Summaries")
            if os.path.isfile(MARKDOWN_CSV):
                try:
                    md_df = pd.read_csv(MARKDOWN_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                    all_md = "\n\n---\n\n".join(md_df["markdown_summary"].dropna().tolist())
                    st.download_button(
                        "⬇️ Download All Summaries as Markdown",
                        all_md.encode("utf-8"),
                        file_name="incentive_summaries.md",
                        mime="text/markdown"
                    )
                    with open(MARKDOWN_CSV, "rb") as f:
                        st.download_button(
                            "⬇️ Download Markdown CSV",
                            f,
                            file_name="markdown_summaries.csv"
                        )
                    with _scrollable_container(height=600):
                        for _, row in md_df.iterrows():
                            st.markdown(row["markdown_summary"])
                            st.markdown("---")
                except Exception as e:
                    st.error(f"Could not load markdown summaries: {e}")
            else:
                st.info("No summaries generated yet.")

        with tab_errors:
            st.subheader("Error Log")
            err_df = _read_csv_safe(ERRORS_CSV)
            if err_df is not None:
                _dataframe(err_df)
                with open(ERRORS_CSV, "rb") as f:
                    st.download_button(
                        "⬇️ Download Error Log",
                        f,
                        file_name="errors.csv"
                    )
            else:
                st.info("No errors logged.")
