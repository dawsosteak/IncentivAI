"""
Discover electric utility incentive URLs by U.S. state via local OpenSERP instance.
Goal: find NEW utility/cooperative/municipal-energy official websites offering
      rebates, grants, or incentive programs — to be merged into the DSIRE URL database.

Results are saved to an Excel file (one row per search result URL).

Usage:
    python energy_search.py                        # interactive
    python energy_search.py Texas                  # single state
    python energy_search.py "New York" California  # multiple states, appends to same file
"""

import re
import sys
import time
import datetime
import requests
from typing import List, Dict, Any
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not found — run: pip install openpyxl")

# ── Configuration ─────────────────────────────────────────────────────────────
OPENSERP_URL  = "http://localhost:7000"
SEARCH_ENGINE = "google"
NUM_RESULTS   = 8       # more results per query to maximize new URL discovery
COUNTRY       = "us"
LANGUAGE      = "EN"
OUTPUT_FILE   = "utility_urls_discovered.xlsx"

# ── Search topics ─────────────────────────────────────────────────────────────
# Goal: surface UTILITY OFFICIAL WEBSITES (not news/aggregators).
# Each phrase + state name => Google query.
# Designed to match domain types seen in DSIRE:
#   .coop, municipalelectric.*, *electric.com, energy.*.gov, pud.*.*, etc.
# ── Search topics ─────────────────────────────────────────────────────────────
# Derived from EIA-861 entity naming patterns:
#   - Cooperative / Coop / Electric Assn  (~294 entities)
#   - City of / Municipal Electric        (~595 entities)
#   - Public Utility District / PUD       (~37 political subdivision entities)
#   - Investor Owned / Light & Power      (~60 entities)
# Goal: surface OFFICIAL UTILITY WEBSITES, not news or aggregators.

TOPICS = [
    # ── By utility type (maps to EIA-861 Ownership column) ──
    "electric cooperative rebate incentive program",        # Cooperative
    "electric coop energy efficiency rebate apply",         # Coop (abbrev.)
    "electric association rebate program",                  # Electric Assn
    "municipal electric utility rebate incentive",          # Municipal
    "city electric utility energy rebate program",          # City of ...
    "public utility district rebate incentive program",     # PUD / Political Subdivision
    "rural electric cooperative incentive apply",           # Rural Coop
    "investor owned utility energy efficiency rebate",      # Investor Owned
    "light and power company rebate program",               # Light & Power Co.
    "county electric cooperative rebate program",           # County Electric

    # ── By device/program type (DSIRE program categories) ──
    "electric utility solar rebate apply",
    "electric utility heat pump rebate program",
    "electric utility EV charger rebate apply",
    "electric utility smart thermostat rebate",
    "electric utility battery storage incentive",
    "electric utility weatherization rebate low income",
    "electric utility net metering program",
    "electric utility on-bill financing program",
    "electric utility demand response incentive",
    "electric utility energy efficiency rebate commercial",
]

# ── Domain filter ─────────────────────────────────────────────────────────────
# Blocklist: aggregators, news, advocacy, govt-non-utility, short links
# A result is KEPT if it does NOT match any of these patterns.
DOMAIN_BLOCKLIST = re.compile(
    r"(dsire|energysage|energystar|epa\.gov|energy\.gov$"
    r"|nrel\.gov|eia\.gov|wikipedia|energycoalition"
    r"|nrdc\.org|sierraclub|greentechmedia|pv-magazine"
    r"|forbes|bloomberg|reuters|apnews|cnn\.com|nytimes"
    r"|bit\.ly|tinyurl|t\.co"
    r"|bcap-ocean|cleanairfleets|bcapcodes"
    r"|mgaleg|comptroller|sos\."          # legislative / secretary of state
    r")",
    re.IGNORECASE,
)

def is_utility_url(url: str) -> bool:
    """
    Heuristic: keep URLs that look like utility/government-energy official sites.
    Drops aggregators, news sites, and known non-utility domains.
    """
    try:
        domain = urlparse(url).netloc.lower()
    except Exception:
        return False
    if DOMAIN_BLOCKLIST.search(domain):
        return False
    return True

# ── Excel helpers ─────────────────────────────────────────────────────────────
COL_WIDTHS = {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}

def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _setup_sheet(ws):
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

def _write_header(ws):
    headers = ["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"]
    border = _thin_border()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 20

def _append_row(ws, state, query, url, title, desc, discovered_at):
    border = _thin_border()
    row_data = [state, query, url, title, desc, discovered_at]
    r = ws.max_row + 1
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border    = border
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
        if c == 1:   # State — highlight
            cell.fill = PatternFill("solid", start_color="D6E4F0")
            cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    ws.row_dimensions[r].height = 45

def _get_or_create_workbook(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws  = wb.active
        ws.title = "Discovered URLs"
        _setup_sheet(ws)
        _write_header(ws)
    return wb, ws

# ── Core search ───────────────────────────────────────────────────────────────
def search(query: str, limit: int = NUM_RESULTS) -> List[Dict[str, Any]]:
    url    = f"{OPENSERP_URL}/{SEARCH_ENGINE}/search"
    params = {"text": query, "limit": limit, "gl": COUNTRY, "lang": LANGUAGE}
    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  ❌ Search error: {e}")
        return []

# ── Per-state search ──────────────────────────────────────────────────────────
def load_existing_domains(path: str) -> set:
    """Load existing URL database and return a set of bare domains for deduplication."""
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        domains = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.startswith("http"):
                    d = urlparse(cell.strip()).netloc.lower().lstrip("www.")
                    if d:
                        domains.add(d)
        wb.close()
        print(f"📂 Loaded {len(domains)} existing domains from {path}")
        return domains
    except FileNotFoundError:
        print(f"⚠️  No existing database found at '{path}' — all results will be kept.")
        return set()

# Loaded once at startup, shared across all states
EXISTING_DOMAINS: set = set()

def run_state(state: str) -> None:
    print(f"\n🔍 Discovering utility URLs for: {state}")
    wb, ws = (_get_or_create_workbook(OUTPUT_FILE) if EXCEL_AVAILABLE else (None, None))

    total_found = 0
    total_kept  = 0

    for topic in TOPICS:
        query       = f"{topic} {state}"
        discovered  = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"\n  📌 {query}")

        results = search(query)
        if not results:
            print("     No results.")
            time.sleep(1)
            continue

        kept = []
        for item in results:
            url    = item.get("url", "")
            domain = urlparse(url).netloc.lower().lstrip("www.")
            if not is_utility_url(url):
                print(f"     u26d4 blocked : {url}")
            elif domain in EXISTING_DOMAINS:
                print(f"     u23edufe0f  exists  : {domain}")
            else:
                kept.append(item)

        total_found += len(results)
        total_kept  += len(kept)

        if not kept:
            print("     No utility URLs after filtering.")
        else:
            for item in kept:
                url   = item.get("url", "")
                title = item.get("title", "")
                desc  = item.get("description", "")
                print(f"     ✅ {title[:60]}")
                print(f"        {url}")
                if EXCEL_AVAILABLE and ws is not None:
                    _append_row(ws, state, query, url, title, desc, discovered)

        time.sleep(1)

    if EXCEL_AVAILABLE and wb is not None:
        wb.save(OUTPUT_FILE)

    print(f"\n  📊 {state}: {total_kept}/{total_found} URLs kept → {OUTPUT_FILE}")

# ── Entry points ──────────────────────────────────────────────────────────────
VALID_STATES = {
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
}

def _interactive():
    print(f"🔍 Utility URL Discovery\nOpenSERP: {OPENSERP_URL} | Engine: {SEARCH_ENGINE}\n")
    while True:
        state = input("🗺️  Enter a U.S. state: ").strip().title()
        if state in VALID_STATES:
            break
        print(f"  ⚠️  '{state}' not recognized. Try again.")
    run_state(state)

def _cli(states: list):
    for raw in states:
        state = raw.strip().title()
        if state not in VALID_STATES:
            print(f"⚠️  Skipping '{state}' — not a recognized U.S. state.")
            continue
        print(f"\n{'='*52}\n  🗺️  {state}\n{'='*52}")
        run_state(state)

DATABASE_FILE = "Relevant_URLs.xlsx"   # existing URL database for deduplication

if __name__ == "__main__":
    # Pre-load existing domains so every search skips already-known sites
    EXISTING_DOMAINS = load_existing_domains(DATABASE_FILE)
    print()
    if len(sys.argv) > 1:
        _cli(sys.argv[1:])
    else:
        _interactive()

DATABASE_FILE = "Relevant_URLs.xlsx"
