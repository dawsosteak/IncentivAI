import asyncio
import contextlib
import io
import json
import os
import re
import subprocess
import sys
import time
import zipfile
from datetime import date, datetime
from urllib.parse import urlparse

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
APP_VERSION = "compat-2026-05-07-2"
DEFAULT_URL_COLUMN_NAMES = ("url", "urls", "link", "links", "website", "websites")
DEFAULT_MODEL_OPTIONS = {
    "ollama": ["llama3.2", "qwen2.5:14b-instruct", "mistral"],
    "openai": ["gpt-4.1", "gpt-4.1-mini", "gpt-4o-mini"],
    "uw_ssec": ['gemma-4-31b', 'olmo-3.1-32b', 'gpt-5.4-pro', 'gpt-oss-120b', 'gpt-5.3-codex', 'devstral-small', 'gpt-5.4-mini'],
    "anthropic": ["claude-sonnet-4-0", "claude-3-5-sonnet-latest"],
    "google": ["gemini-2.5-pro", "gemini-2.5-flash"],
}

VALID_STATES = [
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
]

TOPICS = [
    "electric cooperative rebate incentive program",
    "electric coop energy efficiency rebate apply",
    "electric association rebate program",
    "municipal electric utility rebate incentive",
    "city electric utility energy rebate program",
    "public utility district rebate incentive program",
    "rural electric cooperative incentive apply",
    "investor owned utility energy efficiency rebate",
    "light and power company rebate program",
    "county electric cooperative rebate program",
    "electric utility solar rebate apply",
    "electric utility heat pump rebate program",
    "electric utility EV charger rebate apply",
    "electric utility smart thermostat rebate",
    "electric utility battery storage incentive",
    "electric utility weatherization rebate low income",
    "electric utility net metering program",
    "electric utility on-bill financing program",
    "electric utility demand response incentive",
    "electric utility energy efficiency rebate commercial",
]

DOMAIN_BLOCKLIST = re.compile(
    r"(dsire|energysage|energystar|epa\.gov|energy\.gov$"
    r"|nrel\.gov|eia\.gov|wikipedia|energycoalition"
    r"|nrdc\.org|sierraclub|greentechmedia|pv-magazine"
    r"|forbes|bloomberg|reuters|apnews|cnn\.com|nytimes"
    r"|bit\.ly|tinyurl|t\.co"
    r"|bcap-ocean|cleanairfleets|bcapcodes"
    r"|mgaleg|comptroller|sos\.)",
    re.IGNORECASE,
)


class StreamlitPrintTee(io.TextIOBase):
    def __init__(self, stream, on_write):
        self.stream = stream
        self.on_write = on_write

    @property
    def encoding(self):
        return getattr(self.stream, "encoding", "utf-8")

    def writable(self):
        return True

    def write(self, text):
        self.stream.write(text)
        self.stream.flush()
        if text:
            self.on_write(text)
        return len(text)

    def flush(self):
        self.stream.flush()


class PipelineRunError(Exception):
    def __init__(self, message, log_text):
        super().__init__(message)
        self.log_text = log_text


def _find_default_url_column(columns):
    normalized = {str(col).strip().lower(): col for col in columns}
    for name in DEFAULT_URL_COLUMN_NAMES:
        if name in normalized:
            return normalized[name]
    for col in columns:
        lowered = str(col).strip().lower()
        if "url" in lowered or "link" in lowered:
            return col
    return columns[0] if len(columns) else None


def _normalize_url(value):
    if pd.isna(value):
        return ""
    url = str(value).strip()
    if not url:
        return ""
    parsed = urlparse(url)
    if parsed.scheme and parsed.netloc:
        return url
    if parsed.netloc:
        return f"https:{url}"
    if "." in url and " " not in url:
        return f"https://{url}"
    return ""


def _default_models_for_provider(provider):
    return DEFAULT_MODEL_OPTIONS.get((provider or "").lower(), ["llama3.2"])


@st.cache_data(show_spinner=False)
def _get_ollama_models():
    try:
        result = subprocess.run(["ollama", "list"], capture_output=True, text=True, check=True)
    except Exception:
        return []

    models = []
    lines = result.stdout.splitlines()
    for line in lines[1:]:
        parts = line.split()
        if parts:
            models.append(parts[0])
    return models


def _run_pipeline_for_url(
    url,
    use_deep_crawl,
    provider,
    model_name,
    truncation_length,
    log_placeholder=None,
):
    from test_single_link import (
        analyze_scraped_files,
        filter_analysis_results,
        scrape_single_link,
    )

    log_chunks = []

    def append_log(text):
        log_chunks.append(text)
        if log_placeholder is not None:
            log_placeholder.text("".join(log_chunks)[-12000:])

    stdout_tee = StreamlitPrintTee(sys.stdout, append_log)
    stderr_tee = StreamlitPrintTee(sys.stderr, append_log)

    try:
        with contextlib.redirect_stdout(stdout_tee), contextlib.redirect_stderr(stderr_tee):
            scraped_files = asyncio.run(
                scrape_single_link(
                    url,
                    use_deep_crawl=use_deep_crawl,
                    truncation_length=truncation_length,
                )
            )
            analysis_files = analyze_scraped_files(scraped_files, provider=provider, model_name=model_name)
            filter_analysis_results(analysis_files, provider=provider, model_name=model_name)
    except Exception as exc:
        log_text = "".join(log_chunks)
        raise PipelineRunError(str(exc), log_text) from exc

    final_files = []
    for analysis_file in analysis_files:
        base_name = os.path.basename(analysis_file).replace("_analysis.md", "")
        final_file = os.path.join(
            os.path.dirname(analysis_file),
            f"{base_name}_FINAL_rebates.md",
        )
        if os.path.exists(final_file):
            final_files.append(final_file)

    return scraped_files, analysis_files, final_files, "".join(log_chunks)


def _build_zip(filepaths):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in filepaths:
            zf.write(path, arcname=os.path.basename(path))
    buffer.seek(0)
    return buffer


def _markdown_table(rows, columns):
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows:
        values = []
        for column in columns:
            value = str(row.get(column, ""))
            value = value.replace("|", "\\|").replace("\n", " ")
            values.append(value)
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def _caption(text):
    caption = getattr(st, "caption", None)
    if callable(caption):
        caption(text)
    else:
        st.write(text)


def _divider():
    divider = getattr(st, "divider", None)
    if callable(divider):
        divider()
    else:
        st.markdown("---")


def _expander(label, expanded=False):
    expander = getattr(st, "expander", None)
    if callable(expander):
        return expander(label, expanded=expanded)
    st.write(f"**{label}**")
    return contextlib.nullcontext()


def _download_button(label, data, file_name, mime):
    download_button = getattr(st, "download_button", None)
    if callable(download_button):
        download_button(label, data, file_name=file_name, mime=mime)
    else:
        st.info(f"Download buttons are not available in this Streamlit version: {file_name}")


# Discovery helpers, adapted from Searching/energy_search.py without changing it.
def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def is_utility_url(url):
    try:
        domain = urlparse(url).netloc.lower()
    except Exception:
        return False
    return not DOMAIN_BLOCKLIST.search(domain)


def _xl_get_or_create(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Discovered URLs"
        for column, width in {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"
        border = _thin_border()
        headers = ["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"]
        for column_index, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=column_index, value=header)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", start_color="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 20
    return wb, ws


def _xl_append_row(ws, state, query, url, title, description, discovered_at):
    border = _thin_border()
    row_index = ws.max_row + 1
    values = [state, query, url, title, description, discovered_at]
    for column_index, value in enumerate(values, 1):
        cell = ws.cell(row=row_index, column=column_index, value=value)
        cell.border = border
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=(column_index in (2, 3, 4, 5)))
        if column_index == 1:
            cell.fill = PatternFill("solid", start_color="D6E4F0")
            cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
    ws.row_dimensions[row_index].height = 45


def _extract_domain(url):
    try:
        return urlparse(url.strip()).netloc.lower().lstrip("www.")
    except Exception:
        return ""


def _load_existing_domains(source):
    try:
        wb = load_workbook(source, read_only=True)
    except FileNotFoundError:
        return set()

    ws = wb.active
    domains = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and cell.startswith("http"):
                domain = _extract_domain(cell)
                if domain:
                    domains.add(domain)
    wb.close()
    return domains


def _search_openserp(query, openserp_url, engine, limit=8):
    search_url = f"{openserp_url.rstrip('/')}/{engine}/search"
    params = {"text": query, "limit": limit, "gl": "us", "lang": "EN"}
    try:
        response = requests.get(search_url, params=params, timeout=15)
        response.raise_for_status()
        return response.json()
    except Exception:
        return []


# Merge helpers, adapted from Searching/merge_urls.py without changing it.
def _load_existing_urls(source):
    wb = load_workbook(source, read_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and cell.startswith("http"):
                urls.append(cell.strip())
    wb.close()
    return urls


def _load_discovered_file(source):
    wb = load_workbook(source, read_only=True)
    ws = wb.active
    rows = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(cell).strip() if cell else "" for cell in row]
            continue
        if not any(row):
            continue
        record = dict(zip(headers, row))
        url = str(record.get("URL", "") or "").strip()
        if url.startswith("http"):
            rows.append(record)
    wb.close()
    return rows


def _build_merged_workbook(existing_urls, new_rows):
    wb = Workbook()
    border = _thin_border()

    ws_all = wb.active
    ws_all.title = "All URLs"
    ws_all.column_dimensions["A"].width = 80
    ws_all.freeze_panes = "A2"

    header = ws_all.cell(row=1, column=1, value="Program Source URLs")
    header.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header.fill = PatternFill("solid", start_color="1F4E79")
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.border = border
    ws_all.row_dimensions[1].height = 20

    for row_index, url in enumerate(existing_urls, 2):
        cell = ws_all.cell(row=row_index, column=1, value=url)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top")
        cell.border = border

    separator_row = len(existing_urls) + 2
    separator = ws_all.cell(row=separator_row, column=1, value=f"-- NEW URLS ADDED {date.today()} --")
    separator.font = Font(name="Arial", bold=True, color="2E75B6", size=10)
    separator.fill = PatternFill("solid", start_color="D6E4F0")

    for row_index, row in enumerate(new_rows, separator_row + 1):
        cell = ws_all.cell(row=row_index, column=1, value=str(row.get("URL", "")).strip())
        cell.font = Font(name="Arial", size=10, color="2E75B6")
        cell.alignment = Alignment(vertical="top")
        cell.border = border
        cell.fill = PatternFill("solid", start_color="EBF3FB")

    ws_new = wb.create_sheet("New URLs")
    for column, width in {"A": 16, "B": 35, "C": 65, "D": 45, "E": 65, "F": 22}.items():
        ws_new.column_dimensions[column].width = width
    ws_new.freeze_panes = "A2"

    headers = ["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"]
    for column_index, header_text in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=column_index, value=header_text)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws_new.row_dimensions[1].height = 20

    for row_index, row in enumerate(new_rows, 2):
        values = [
            row.get("State", ""),
            row.get("Search Query", ""),
            row.get("URL", ""),
            row.get("Page Title", ""),
            row.get("Description", ""),
            row.get("Discovered At", ""),
        ]
        for column_index, value in enumerate(values, 1):
            cell = ws_new.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(column_index in (2, 3, 4, 5)))
            if column_index == 1:
                cell.fill = PatternFill("solid", start_color="D6E4F0")
                cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        ws_new.row_dimensions[row_index].height = 45

    return wb


try:
    st.set_page_config(page_title="Incentive Pipeline", layout="wide")
except TypeError:
    st.set_page_config(page_title="Incentive Pipeline")

st.title("Incentive Pipeline")
_caption("Run the existing pipeline, discover new utility URLs, or merge URL databases.")
_caption(f"App version: {APP_VERSION}")

input_mode = st.radio(
    "Mode",
    ["Excel upload", "Single URL", "Scraped markdown directory", "URL Discovery", "Merge Database"],
    horizontal=True,
)
urls = []
scraped_files_dict = {}  # Maps URL to scraped files for directory mode

if input_mode == "Excel upload":
    uploaded_file = st.file_uploader("Excel file", type=["xlsx"])

    if uploaded_file:
        workbook = pd.ExcelFile(uploaded_file)
        sheet_name = st.selectbox("Sheet", workbook.sheet_names)
        df = pd.read_excel(workbook, sheet_name=sheet_name)

        if df.empty:
            st.warning("This sheet is empty.")
            st.stop()

        default_column = _find_default_url_column(df.columns)
        column_options = list(df.columns)
        default_index = column_options.index(default_column) if default_column in column_options else 0
        url_column = st.selectbox("URL column", column_options, index=default_index)

        urls = [_normalize_url(value) for value in df[url_column]]
        urls = [url for url in urls if url]
        urls = list(dict.fromkeys(urls))

        st.write(f"Found {len(urls)} valid unique URL(s).")
        with _expander("Preview URLs"):
            preview_rows = [{"url": url} for url in urls[:50]]
            st.markdown(_markdown_table(preview_rows, ["url"]))
            if len(urls) > 50:
                _caption(f"Showing the first 50 of {len(urls)} URLs.")
    else:
        st.info("Upload an Excel workbook to begin.")
elif input_mode == "Single URL":
    single_url = st.text_input("URL", placeholder="https://www.example.com/rebates")
    normalized_url = _normalize_url(single_url)

    if single_url and normalized_url:
        urls = [normalized_url]
        st.write(f"Ready to run: {normalized_url}")
    elif single_url:
        st.warning("Enter a valid URL, such as https://www.example.com/rebates.")
    else:
        st.info("Enter a URL to begin.")

elif input_mode == "Scraped markdown directory":
    directory = st.text_input("Directory path", placeholder=os.path.join(BASE_DIR, "scraped_markdown"))
    if directory and os.path.isdir(directory):
        md_files = [
            os.path.join(directory, f)
            for f in os.listdir(directory)
            if f.endswith(".md")
        ]
        if md_files:
            for md_file in md_files:
                filename = os.path.basename(md_file)
                url_part = filename.split("_scraped.md")[0]
                url = url_part.replace("_", "/").replace("~", ":")
                scraped_files_dict[url] = [md_file]
            urls = list(scraped_files_dict.keys())
            st.write(f"Found {len(urls)} URLs with scraped markdown files in the directory.")
        else:
            st.warning("No scraped markdown files found in the directory.")
    elif directory:
        st.warning("Enter a valid directory path that contains scraped markdown files.")
    else:
        st.info("Enter a directory path to begin.")

elif input_mode == "URL Discovery":
    st.subheader("Discover New Utility URLs")
    _caption("Search OpenSERP for electric utility/cooperative websites by state and skip domains already in your database.")

    searching_dir = os.path.join(BASE_DIR, "Searching")
    default_db_path = os.path.join(searching_dir, "Relevant_URLs.xlsx")
    default_output_path = os.path.join(searching_dir, "utility_urls_discovered.xlsx")

    col_left, col_right = st.columns(2)
    with col_left:
        selected_states = st.multiselect("States to search", VALID_STATES, default=["Texas"])
        openserp_url = st.text_input("OpenSERP URL", value="http://localhost:7070")
        engine = st.selectbox("Search engine", ["google", "bing", "duckduckgo"], index=0)
    with col_right:
        num_results = st.slider("Results per query", min_value=3, max_value=15, value=8)
        output_file = st.text_input("Output file path", value=default_output_path)
        db_file = st.file_uploader(
            "Existing URL database for deduplication (optional)",
            type=["xlsx"],
            key="discovery_db_upload",
        )
        db_status = "found" if os.path.exists(default_db_path) else "not found"
        _caption(f"Default database: {default_db_path} ({db_status})")

    st.markdown(f"**Search topics:** {len(TOPICS)} queries per state")
    with _expander("View or edit topics"):
        topics_text = st.text_area("One topic per line", value="\n".join(TOPICS), height=300)
        active_topics = [topic.strip() for topic in topics_text.splitlines() if topic.strip()]

    run_disabled = not selected_states or not active_topics or not output_file.strip()
    if st.button("Run Discovery", disabled=run_disabled):
        existing_domains = set()
        if db_file:
            existing_domains = _load_existing_domains(io.BytesIO(db_file.read()))
            st.info(f"Loaded {len(existing_domains)} existing domains from uploaded file.")
        elif os.path.exists(default_db_path):
            existing_domains = _load_existing_domains(default_db_path)
            st.info(f"Auto-loaded {len(existing_domains)} existing domains from Searching/Relevant_URLs.xlsx.")

        output_path = output_file.strip()
        total_queries = len(selected_states) * len(active_topics)
        progress = st.progress(0)
        status = st.empty()
        log_area = st.empty()
        log_lines = []
        discovered_rows = []
        query_count = 0

        wb, ws = _xl_get_or_create(output_path)

        for state in selected_states:
            for topic in active_topics:
                query_count += 1
                query = f"{topic} {state}"
                progress.progress(query_count / total_queries)
                status.write(f"Searching: {query}")

                results = _search_openserp(query, openserp_url, engine, limit=num_results)
                discovered_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                for item in results:
                    result_url = item.get("url", "")
                    domain = _extract_domain(result_url)
                    if not result_url or not domain:
                        log_lines.append(f"skipped empty result for: {query}")
                    elif not is_utility_url(result_url):
                        log_lines.append(f"blocked: {result_url}")
                    elif domain in existing_domains:
                        log_lines.append(f"exists: {domain}")
                    else:
                        existing_domains.add(domain)
                        title = item.get("title", "")
                        description = item.get("description", "")
                        row = {
                            "State": state,
                            "Search Query": query,
                            "URL": result_url,
                            "Page Title": title,
                            "Description": description,
                            "Discovered At": discovered_at,
                        }
                        _xl_append_row(ws, state, query, result_url, title, description, discovered_at)
                        discovered_rows.append(row)
                        log_lines.append(f"new: {result_url}")

                log_area.text("\n".join(log_lines[-60:]) or "No results yet.")
                time.sleep(0.8)

        wb.save(output_path)
        progress.progress(100)
        status.write("Discovery complete.")
        st.success(f"Found {len(discovered_rows)} new utility URL(s) across {len(selected_states)} state(s).")

        if discovered_rows:
            st.subheader("Discovered URLs")
            st.markdown(_markdown_table(discovered_rows, ["State", "URL", "Page Title"]))
            with open(output_path, "rb") as output_handle:
                _download_button(
                    "Download discovered URLs Excel",
                    output_handle.read(),
                    file_name=os.path.basename(output_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

elif input_mode == "Merge Database":
    st.subheader("Merge Discovered URLs into Database")
    _caption("Domain-level deduplication for merging utility_urls_discovered.xlsx into Relevant_URLs.xlsx.")

    searching_dir = os.path.join(BASE_DIR, "Searching")
    auto_db_path = os.path.join(searching_dir, "Relevant_URLs.xlsx")
    auto_discovered_path = os.path.join(searching_dir, "utility_urls_discovered.xlsx")
    db_status = "found" if os.path.exists(auto_db_path) else "not found"
    discovered_status = "found" if os.path.exists(auto_discovered_path) else "not found"
    st.info(f"Database: `{auto_db_path}` ({db_status})\n\nDiscovered: `{auto_discovered_path}` ({discovered_status})")

    col_left, col_right = st.columns(2)
    with col_left:
        db_upload = st.file_uploader("Override existing URL database", type=["xlsx"], key="merge_db_upload")
    with col_right:
        discovered_upload = st.file_uploader("Override discovered URLs file", type=["xlsx"], key="merge_discovered_upload")

    def _resolve_workbook_source(upload, auto_path, label):
        if upload:
            return io.BytesIO(upload.read()), f"uploaded {label}"
        if os.path.exists(auto_path):
            return auto_path, os.path.join("Searching", os.path.basename(auto_path))
        return None, None

    db_source, db_source_label = _resolve_workbook_source(db_upload, auto_db_path, "database")
    discovered_source, discovered_source_label = _resolve_workbook_source(
        discovered_upload,
        auto_discovered_path,
        "discovered URLs",
    )

    if db_source and discovered_source:
        existing_urls = _load_existing_urls(db_source)
        existing_domains = {_extract_domain(url) for url in existing_urls if _extract_domain(url)}
        discovered_rows = _load_discovered_file(discovered_source)
        st.caption(f"Loaded database from {db_source_label}; discovered URLs from {discovered_source_label}.")

        st.write(f"**Database:** {len(existing_urls)} URLs / {len(existing_domains)} unique domains")
        st.write(f"**Discovered:** {len(discovered_rows)} rows")

        seen_domains = set(existing_domains)
        new_rows = []
        skipped = 0
        for row in discovered_rows:
            result_url = str(row.get("URL", "") or "").strip()
            domain = _extract_domain(result_url)
            if not domain:
                continue
            if domain in seen_domains:
                skipped += 1
            else:
                seen_domains.add(domain)
                new_rows.append(row)

        st.write(f"**After dedup:** {len(new_rows)} new URLs, {skipped} skipped duplicate domain(s)")

        if new_rows:
            with _expander("Preview new URLs"):
                st.markdown(_markdown_table(new_rows[:30], ["State", "URL", "Page Title"]))
                if len(new_rows) > 30:
                    _caption(f"Showing first 30 of {len(new_rows)} URLs.")

        if st.button("Merge and Download", disabled=not new_rows):
            merged_workbook = _build_merged_workbook(existing_urls, new_rows)
            buffer = io.BytesIO()
            merged_workbook.save(buffer)
            buffer.seek(0)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success(f"Merged {len(existing_urls)} existing URL(s) with {len(new_rows)} new URL(s).")
            _download_button(
                "Download merged database",
                buffer.read(),
                file_name=f"Relevant_URLs_merged_{stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Auto-detected files were not found. Upload both workbooks above to merge.")

if urls and input_mode in ("Excel upload", "Single URL", "Scraped markdown directory"):
    _divider()
    use_deep_crawl = st.checkbox("Deep crawl", value=True)
    provider = st.selectbox("LLM provider", ["ollama", "uw_ssec", "openai", "anthropic", "google"], index=0)
    model_choices = _get_ollama_models() if provider == "ollama" else _default_models_for_provider(provider)
    if provider == "ollama" and not model_choices:
        model_choices = _default_models_for_provider(provider)
    model_name = st.selectbox("Model", model_choices, index=0)
    custom_model = st.text_input("Custom model name (optional)", value="")
    model_name = custom_model.strip() or model_name
    truncation_length = st.number_input(
        "Max scrape length",
        min_value=10000,
        max_value=500000,
        value=150000,
        step=10000,
    )

    if st.button("Run Pipeline"):
        progress = st.progress(0)
        status = st.empty()
        with _expander("Live pipeline log", expanded=True):
            log_placeholder = st.empty()
        rows = []
        final_files = []
        log_text = ""

        for index, url in enumerate(urls, start=1):
            status.write(f"Running {index} of {len(urls)}: {url}")
            try:
                # Skip scraping if using directory mode
                if input_mode == "Scraped markdown directory":
                    scraped = scraped_files_dict.get(url, [])
                    log_text = f"Using pre-scraped files from directory: {scraped}"
                else:
                    scraped, _, _, _ = _run_pipeline_for_url(
                        url=url,
                        use_deep_crawl=use_deep_crawl,
                        provider=provider,
                        model_name=model_name,
                        truncation_length=int(truncation_length),
                        log_placeholder=log_placeholder,
                    )
                
                from test_single_link import analyze_scraped_files, filter_analysis_results
                
                analyzed = analyze_scraped_files(scraped, provider=provider, model_name=model_name)
                filter_analysis_results(analyzed, provider=provider, model_name=model_name)
                
                finals = []
                for analysis_file in analyzed:
                    base_name = os.path.basename(analysis_file).replace("_analysis.md", "")
                    final_file = os.path.join(
                        os.path.dirname(analysis_file),
                        f"{base_name}_FINAL_rebates.md",
                    )
                    if os.path.exists(final_file):
                        finals.append(final_file)
                
                final_files.extend(finals)
                rows.append(
                    {
                        "url": url,
                        "status": "done",
                        "scraped_files": len(scraped),
                        "analysis_files": len(analyzed),
                        "final_files": len(finals),
                        "error": "",
                    }
                )
            except Exception as exc:
                log_text = getattr(exc, "log_text", "")
                log_text = f"{log_text}\nERROR: {exc}" if log_text else str(exc)
                rows.append(
                    {
                        "url": url,
                        "status": "error",
                        "scraped_files": 0,
                        "analysis_files": 0,
                        "final_files": 0,
                        "error": str(exc),
                    }
                )

            progress.progress(int(index / len(urls) * 100))
            log_placeholder.text(log_text[-12000:] or "No log output yet.")

        status.write("Run complete.")
        results_df = pd.DataFrame(rows)
        st.subheader("Run Summary")
        st.markdown(_markdown_table(rows, list(results_df.columns)))

        csv_bytes = results_df.to_csv(index=False).encode("utf-8")
        _download_button(
            "Download run summary CSV",
            csv_bytes,
            file_name="single_link_tester_run_summary.csv",
            mime="text/csv",
        )

        unique_final_files = list(dict.fromkeys(final_files))
        if unique_final_files:
            zip_buffer = _build_zip(unique_final_files)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            _download_button(
                "Download final rebate markdown zip",
                zip_buffer,
                file_name=f"final_rebates_{stamp}.zip",
                mime="application/zip",
            )

            st.subheader("Final Results")
            for path in unique_final_files:
                with _expander(os.path.basename(path)):
                    with open(path, "r", encoding="utf-8") as f:
                        st.markdown(f.read())
        else:
            st.info("No final rebate markdown files were produced.")
