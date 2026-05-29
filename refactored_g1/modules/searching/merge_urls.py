"""
Merge newly discovered utility URLs into the existing Relevant_URLs.xlsx database.
Deduplication is done at the DOMAIN level — same domain = already exists, skip.

Usage:
    python merge_urls.py
    python merge_urls.py --discovered utility_urls_discovered.xlsx --database Relevant_URLs.xlsx
"""

import argparse
import os
import datetime
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl not found — run: pip install openpyxl")
    raise

# ── Defaults ──────────────────────────────────────────────────────────────────
DEFAULT_DISCOVERED = "utility_urls_discovered.xlsx"
DEFAULT_DATABASE   = "Relevant_URLs.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────────
def extract_domain(url: str) -> str:
    """Normalize URL to bare domain, stripping www. prefix."""
    try:
        domain = urlparse(url.strip()).netloc.lower()
        return domain.lstrip("www.")
    except Exception:
        return ""

def load_existing_domains(path: str) -> set:
    """Load all URLs from the database file and return a set of their domains."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    domains = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and cell.startswith("http"):
                d = extract_domain(cell)
                if d:
                    domains.add(d)
    wb.close()
    return domains

def load_existing_urls(path: str) -> list:
    """Load all existing URLs as a plain list."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and cell.startswith("http"):
                urls.append(cell.strip())
    wb.close()
    return urls

def load_discovered(path: str) -> list[dict]:
    """
    Load rows from utility_urls_discovered.xlsx.
    Expected columns: State / Search Query / URL / Page Title / Description / Discovered At
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        record = dict(zip(headers, row))
        url = str(record.get("URL", "") or "").strip()
        if url.startswith("http"):
            rows.append(record)
    wb.close()
    return rows

# ── Excel output ──────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def build_merged_workbook(existing_urls: list, new_rows: list) -> Workbook:
    """
    Sheet 1 — "All URLs"    : original URLs + new URLs (URL column only, matches original format)
    Sheet 2 — "New URLs"    : new URLs with full metadata for review
    """
    wb = Workbook()

    # ── Sheet 1: All URLs ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All URLs"
    ws1.column_dimensions["A"].width = 80
    ws1.freeze_panes = "A2"

    # header
    h = ws1.cell(row=1, column=1, value="Program Source URLs")
    h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    h.fill      = PatternFill("solid", start_color="1F4E79")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border    = _thin_border()
    ws1.row_dimensions[1].height = 20

    for r, url in enumerate(existing_urls, 2):
        cell = ws1.cell(row=r, column=1, value=url)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top")
        cell.border    = _thin_border()

    separator_row = len(existing_urls) + 2
    sep = ws1.cell(row=separator_row, column=1, value=f"── NEW URLS ADDED {datetime.date.today()} ──")
    sep.font = Font(name="Arial", bold=True, color="2E75B6", size=10)
    sep.fill = PatternFill("solid", start_color="D6E4F0")

    for i, row in enumerate(new_rows, separator_row + 1):
        url  = str(row.get("URL", "")).strip()
        cell = ws1.cell(row=i, column=1, value=url)
        cell.font      = Font(name="Arial", size=10, color="2E75B6")
        cell.alignment = Alignment(vertical="top")
        cell.border    = _thin_border()
        cell.fill      = PatternFill("solid", start_color="EBF3FB")

    # ── Sheet 2: New URLs with metadata ─────────────────────────────────────
    ws2 = wb.create_sheet("New URLs")
    col_widths = {"A": 16, "B": 35, "C": 65, "D": 45, "E": 65, "F": 22}
    for col, w in col_widths.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    headers = ["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"]
    for c, hdr in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=hdr)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws2.row_dimensions[1].height = 20

    for r, row in enumerate(new_rows, 2):
        values = [
            row.get("State", ""),
            row.get("Search Query", ""),
            row.get("URL", ""),
            row.get("Page Title", ""),
            row.get("Description", ""),
            row.get("Discovered At", ""),
        ]
        for c, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.border    = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
            if c == 1:
                cell.fill = PatternFill("solid", start_color="D6E4F0")
                cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        ws2.row_dimensions[r].height = 45

    return wb

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Merge discovered utility URLs into database.")
    parser.add_argument("--discovered", default=DEFAULT_DISCOVERED,
                        help=f"Discovered URLs file (default: {DEFAULT_DISCOVERED})")
    parser.add_argument("--database",   default=DEFAULT_DATABASE,
                        help=f"Existing URL database (default: {DEFAULT_DATABASE})")
    args = parser.parse_args()

    db_path = args.database
    print(f"📂 Database  : {db_path}")
    print(f"📂 Discovered: {args.discovered}\n")

    # Load existing
    print("⏳ Loading existing URL database...")
    existing_domains = load_existing_domains(db_path)
    existing_urls    = load_existing_urls(db_path)
    print(f"   {len(existing_urls)} existing URLs / {len(existing_domains)} unique domains\n")

    # Load discovered
    print("⏳ Loading discovered URLs...")
    discovered = load_discovered(args.discovered)
    print(f"   {len(discovered)} discovered rows\n")

    # Deduplicate
    seen_domains = set(existing_domains)
    new_rows     = []
    skipped      = 0

    for row in discovered:
        url    = str(row.get("URL", "")).strip()
        domain = extract_domain(url)
        if not domain:
            continue
        if domain in seen_domains:
            print(f"  ⛔ duplicate domain: {domain}")
            skipped += 1
        else:
            seen_domains.add(domain)
            new_rows.append(row)
            print(f"  ✅ new: {url}")

    print(f"\n📊 Summary:")
    print(f"   Discovered : {len(discovered)}")
    print(f"   Skipped    : {skipped} (duplicate domain)")
    print(f"   New URLs   : {len(new_rows)}")

    if not new_rows:
        print("\n✅ No new URLs to add. Database is up to date.")
        return

    # Rename old database → Relevant_URLs_YYYYMMDD_HHMMSS.xlsx
    timestamp   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = db_path.replace(".xlsx", f"_backup_{timestamp}.xlsx")
    os.rename(db_path, backup_path)
    print(f"\n📦 Old database renamed → {backup_path}")

    # Save merged file as the new database (same name as old)
    print(f"⏳ Building merged workbook...")
    wb = build_merged_workbook(existing_urls, new_rows)
    wb.save(db_path)
    print(f"✅ Saved → {db_path}  (replaces old database)")
    print(f"   Sheet 1 \'All URLs\'  : {len(existing_urls) + len(new_rows)} total URLs")
    print(f"   Sheet 2 \'New URLs\'  : {len(new_rows)} new entries with metadata")

if __name__ == "__main__":
    main()
