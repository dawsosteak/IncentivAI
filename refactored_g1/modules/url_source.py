import io
import re
import time
import datetime
import requests
import pandas as pd
from urllib.parse import urlparse
from bs4 import BeautifulSoup

# ── Constants ─────────────────────────────────────────────────────────────────

VALID_STATES = [
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming",
]

# Search topics derived from EIA-861 utility naming patterns
DISCOVERY_TOPICS = [
    "electric cooperative rebate incentive program",
    "electric coop energy efficiency rebate apply",
    "electric association rebate program",
    "municipal electric utility rebate incentive",
    "city electric utility energy rebate program",
    "public utility district rebate incentive program",
    "rural electric cooperative incentive apply",
    "investor owned utility energy efficiency rebate",
    "light and power company rebate program",
    "county electric cooperative rebate program",
    "electric utility solar rebate apply",
    "electric utility heat pump rebate program",
    "electric utility EV charger rebate apply",
    "electric utility smart thermostat rebate",
    "electric utility battery storage incentive",
    "electric utility weatherization rebate low income",
    "electric utility net metering program",
    "electric utility on-bill financing program",
    "electric utility demand response incentive",
    "electric utility energy efficiency rebate commercial",
]

# Domains to block — aggregators, news, advocacy, non-utility govt
DOMAIN_BLOCKLIST = re.compile(
    r"(dsire|energysage|energystar|epa\.gov|energy\.gov$"
    r"|nrel\.gov|eia\.gov|wikipedia|energycoalition"
    r"|nrdc\.org|sierraclub|greentechmedia|pv-magazine"
    r"|forbes|bloomberg|reuters|apnews|cnn\.com|nytimes"
    r"|bit\.ly|tinyurl|t\.co"
    r"|bcap-ocean|cleanairfleets|bcapcodes"
    r"|mgaleg|comptroller|sos\.)",
    re.IGNORECASE,
)


# ── Shared helpers ────────────────────────────────────────────────────────────

def _normalize_url(value) -> str:
    """Ensure a URL has a scheme. Returns empty string if invalid."""
    if pd.isna(value):
        return ""
    url = str(value).strip()
    if not url:
        return ""
    parsed = urlparse(url)
    if parsed.scheme and parsed.netloc:
        return url
    if parsed.netloc:
        return f"https:{url}"
    if "." in url and " " not in url:
        return f"https://{url}"
    return ""


def _extract_domain(url: str) -> str:
    """Return bare domain stripped of www. prefix."""
    try:
        return urlparse(url.strip()).netloc.lower().lstrip("www.")
    except Exception:
        return ""


def is_utility_url(url: str) -> bool:
    """Return True if URL looks like a utility official site (not an aggregator/news)."""
    try:
        domain = urlparse(url).netloc.lower()
    except Exception:
        return False
    return not DOMAIN_BLOCKLIST.search(domain)


def _load_existing_domains_from_excel(path_or_file) -> set:
    """
    Load all URLs from an existing Excel database and return their bare domains.
    Used for deduplication during discovery.
    Accepts a file path string or a Streamlit UploadedFile / BytesIO object.
    """
    try:
        from openpyxl import load_workbook
        if hasattr(path_or_file, "read"):
            data = path_or_file.read()
            wb = load_workbook(io.BytesIO(data), read_only=True)
        else:
            wb = load_workbook(path_or_file, read_only=True)
        ws = wb.active
        domains = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and cell.startswith("http"):
                    d = _extract_domain(cell.strip())
                    if d:
                        domains.add(d)
        wb.close()
        return domains
    except Exception:
        return set()


def _build_merged_workbook(existing_urls: list, new_rows: list):
    """
    Build a merged Excel workbook with two sheets:
      Sheet 1 — All URLs: original + new URLs (URL column only)
      Sheet 2 — New URLs: new entries with full metadata

    Args:
        existing_urls: list of URL strings from the existing database
        new_rows:      list of dicts with keys: State, Search Query, URL,
                       Page Title, Description, Discovered At

    Returns:
        openpyxl Workbook object
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()

    # ── Sheet 1: All URLs ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All URLs"
    ws1.column_dimensions["A"].width = 80
    ws1.freeze_panes = "A2"

    h = ws1.cell(row=1, column=1, value="Program Source URLs")
    h.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    h.fill = PatternFill("solid", start_color="1F4E79")
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border = _thin()
    ws1.row_dimensions[1].height = 20

    for r, url in enumerate(existing_urls, 2):
        cell = ws1.cell(row=r, column=1, value=url)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top")
        cell.border = _thin()

    sep_row = len(existing_urls) + 2
    sep = ws1.cell(row=sep_row, column=1, value=f"── NEW URLS ADDED {datetime.date.today()} ──")
    sep.font = Font(name="Arial", bold=True, color="2E75B6", size=10)
    sep.fill = PatternFill("solid", start_color="D6E4F0")

    for i, row in enumerate(new_rows, sep_row + 1):
        cell = ws1.cell(row=i, column=1, value=str(row.get("url", row.get("URL", ""))).strip())
        cell.font = Font(name="Arial", size=10, color="2E75B6")
        cell.alignment = Alignment(vertical="top")
        cell.border = _thin()
        cell.fill = PatternFill("solid", start_color="EBF3FB")

    # ── Sheet 2: New URLs with metadata ──────────────────────────────────────
    ws2 = wb.create_sheet("New URLs")
    for col, w in {"A": 16, "B": 35, "C": 65, "D": 45, "E": 65, "F": 22}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    for c, hdr in enumerate(["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"], 1):
        cell = ws2.cell(row=1, column=c, value=hdr)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
    ws2.row_dimensions[1].height = 20

    for r, row in enumerate(new_rows, 2):
        # Support both lowercase keys (from discovery) and title case (from Excel load)
        values = [
            row.get("state", row.get("State", "")),
            row.get("query", row.get("Search Query", "")),
            row.get("url", row.get("URL", "")),
            row.get("title", row.get("Page Title", "")),
            row.get("description", row.get("Description", "")),
            row.get("discovered_at", row.get("Discovered At", "")),
        ]
        for c, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = _thin()
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
            if c == 1:
                cell.fill = PatternFill("solid", start_color="D6E4F0")
                cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
        ws2.row_dimensions[r].height = 45

    return wb


# ── Mode 1: Excel upload ──────────────────────────────────────────────────────

def _get_urls_from_excel(uploaded_file) -> list[dict]:
    """
    Read URLs from an Excel file.
    Accepts a Streamlit UploadedFile object or a plain file path string.
    Looks for a 'URLs' column and optional 'parent_url' column.

    Returns:
        list of {"url": str, "parent": str | None}
    """
    df = pd.read_excel(uploaded_file)

    # Normalize column names — strip whitespace and lowercase
    df.columns = [c.strip().lower() for c in df.columns]

    if "urls" not in df.columns:
        raise ValueError("Excel file must contain a column named 'URLs'")

    result = []
    for _, row in df.iterrows():
        url = _normalize_url(row["urls"])
        if not url:
            continue

        parent = None
        if "parent_url" in df.columns:
            raw_parent = str(row["parent_url"]).strip()
            if raw_parent and raw_parent.lower() != "nan":
                parent = raw_parent

        result.append({"url": url, "parent": parent})

    # Deduplicate by URL while preserving parent relationship
    seen = set()
    deduped = []
    for entry in result:
        if entry["url"] not in seen:
            seen.add(entry["url"])
            deduped.append(entry)

    return deduped


# ── Mode 2: Auto search by state (DuckDuckGo) ────────────────────────────────

def _get_urls_from_state_search(state: str) -> list[dict]:
    """
    Search DuckDuckGo for electric utility companies in a given state.
    Returns top 10 results as main links.

    Returns:
        list of {"url": str, "parent": None}
    """
    query = f"electric utility companies in {state}"
    search_url = f"https://duckduckgo.com/html/?q={query}"
    try:
        response = requests.get(search_url, timeout=15)
        soup = BeautifulSoup(response.text, "html.parser")
        links = [
            a["href"] for a in soup.find_all("a", href=True)
            if a["href"].startswith("http")
        ]
    except Exception as e:
        raise RuntimeError(f"Auto search failed for state '{state}': {e}")

    seen = set()
    result = []
    for link in links[:10]:
        if link not in seen:
            seen.add(link)
            result.append({"url": link, "parent": None})
    return result


# ── Mode 3: City-based URL discovery via OpenSERP ────────────────────────────

def _search_openserp(query: str, openserp_url: str, engine: str, limit: int = 8) -> list:
    """
    Query a local OpenSERP instance and return raw result list.
    Returns empty list on any failure.
    """
    url = f"{openserp_url}/{engine}/search"
    params = {"text": query, "limit": limit, "gl": "us", "lang": "EN"}
    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return []


def get_urls_from_discovery(
    states: list,
    openserp_url: str = "http://localhost:7070",
    engine: str = "google",
    num_results: int = 8,
    existing_db=None,
    progress_callback=None,
) -> list:
    """
    Discover utility URLs by searching OpenSERP for each state × topic combination.
    Deduplicates against an optional existing URL database.

    Args:
        states:            list of U.S. state names to search
        openserp_url:      base URL of local OpenSERP instance
        engine:            search engine ("google", "bing", "duckduckgo")
        num_results:       results per query
        existing_db:       file path or UploadedFile of existing URL database for dedup
        progress_callback: callable(current, total, url, message) for UI/CLI updates

    Returns:
        list of dicts with keys:
            url, parent, title, description, state, query, discovered_at
    """
    existing_domains = set()
    if existing_db is not None:
        existing_domains = _load_existing_domains_from_excel(existing_db)

    total_queries = len(states) * len(DISCOVERY_TOPICS)
    query_count = 0
    discovered = []
    seen_domains = set(existing_domains)

    for state in states:
        for topic in DISCOVERY_TOPICS:
            query = f"{topic} {state}"
            query_count += 1

            if progress_callback:
                progress_callback(
                    query_count, total_queries,
                    url=query,
                    message=f"({query_count}/{total_queries}) Searching: {query}"
                )

            results = _search_openserp(query, openserp_url, engine, limit=num_results)
            discovered_at = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")

            for item in results:
                url = item.get("url", "")
                domain = _extract_domain(url)

                if not url or not domain:
                    continue
                if not is_utility_url(url):
                    continue
                if domain in seen_domains:
                    continue

                seen_domains.add(domain)
                discovered.append({
                    "url": url,
                    "parent": None,
                    "title": item.get("title", ""),
                    "description": item.get("description", ""),
                    "state": state,
                    "query": query,
                    "discovered_at": discovered_at,
                })

            time.sleep(0.5)  # be polite to OpenSERP

    return discovered


# ── Main entry point ──────────────────────────────────────────────────────────

def get_urls(mode: str, uploaded_file=None, state: str = None) -> list:
    """
    Main URL source router. Returns list of dicts: [{"url": str, "parent": str | None}]

    Modes:
        "Upload Excel"           — read from uploaded Excel file
        "Auto Search Utilities"  — DuckDuckGo search by state name

    Note: "City URL Discovery" is handled directly via get_urls_from_discovery()
    since it requires extra parameters and a progress callback.
    """
    if mode == "Upload Excel":
        return _get_urls_from_excel(uploaded_file)

    if mode == "Auto Search Utilities":
        if not state:
            raise ValueError("State is required for Auto Search mode.")
        return _get_urls_from_state_search(state)

    raise ValueError(f"Unknown mode: '{mode}'. Use 'Upload Excel' or 'Auto Search Utilities'.")
