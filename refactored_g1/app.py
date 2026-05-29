import csv
import io
import os
import datetime
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from main import run_pipeline
from modules.url_source import (
    get_urls_from_discovery,
    VALID_STATES,
    DISCOVERY_TOPICS,
    _load_existing_domains_from_excel,
    _extract_domain,
    _build_merged_workbook,
)
from config import DEFAULT_TEMPERATURE, DEFAULT_TRUNCATION, ERRORS_CSV, MARKDOWN_CSV

# ── Constants ─────────────────────────────────────────────────────────────────
DEFAULT_PROVIDER       = "ollama"
DEFAULT_MODEL          = "qwen2.5:7b"
DEFAULT_TEMP           = DEFAULT_TEMPERATURE
DEFAULT_TRUNCATION_VAL = DEFAULT_TRUNCATION

st.set_page_config(page_title="IncentivAI", layout="wide")
st.title("IncentivAI – Utility Incentive Extractor")

# ── Session state — per-mode keys so switching never erases progress ──────────
_DEFAULTS = {
    "cancelled":            False,
    # Excel mode
    "excel_complete":       False,
    "excel_output_file":    None,
    # Single URL mode
    "single_complete":      False,
    "single_output_file":   None,
    # Markdown mode
    "md_complete":          False,
    "md_output_file":       None,
    "md_input_text":        None,
    "md_input_name":        None,
    # Discovery mode
    "discovery_result":     None,
    "merge_result":         None,
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── Sidebar ───────────────────────────────────────────────────────────────────
st.sidebar.header("Configuration")
mode = st.sidebar.radio(
    "Select Mode",
    ["Upload Excel", "Single URL", "Upload Markdown", "City URL Discovery"]
)

uploaded_file  = None
single_url_in  = None
uploaded_md    = None
temperature       = DEFAULT_TEMP
truncation_length = DEFAULT_TRUNCATION_VAL
provider          = DEFAULT_PROVIDER
model_name        = DEFAULT_MODEL
run_button        = False
cancel_button     = False

if mode == "Upload Excel":
    uploaded_file = st.sidebar.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])

elif mode == "Single URL":
    single_url_in = st.sidebar.text_input("Enter URL to scrape and analyze")

elif mode == "Upload Markdown":
    uploaded_md = st.sidebar.file_uploader("Upload .md or .txt file", type=["md", "txt"])

if mode in ("Upload Excel", "Single URL", "Upload Markdown"):
    st.sidebar.markdown("---")
    temperature       = st.sidebar.number_input("Temperature",       value=DEFAULT_TEMP,           min_value=0.0, max_value=2.0, step=0.1)
    truncation_length = st.sidebar.number_input("Max Scrape Length", value=DEFAULT_TRUNCATION_VAL, min_value=1000, step=1000)
    provider          = st.sidebar.selectbox("LLM Provider", ["ollama", "openai", "anthropic", "google"], index=0)
    model_name        = st.sidebar.text_input("Model Name", value=DEFAULT_MODEL)
    st.sidebar.markdown("---")
    run_button    = st.sidebar.button("▶ Run Extraction", type="primary",  use_container_width=True)
    cancel_button = st.sidebar.button("⏹ Cancel",                          use_container_width=True)

    # Clear only this mode's results
    _pfx = {"Upload Excel": "excel", "Single URL": "single", "Upload Markdown": "md"}[mode]
    if st.sidebar.button("🗑 Clear Results", use_container_width=True):
        st.session_state[f"{_pfx}_complete"]    = False
        st.session_state[f"{_pfx}_output_file"] = None
        st.rerun()

    if cancel_button:
        st.session_state.cancelled = True
        st.sidebar.warning("Stopping after current URL.")


# ── Shared helpers ────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _find_url_column(columns):
    """Best-guess URL column — used only as the default index for the selectbox."""
    _names = {"url", "urls", "link", "links", "website", "websites"}
    for col in columns:
        if str(col).strip().lower() in _names:
            return col
    return None


def _render_markdown_tab(tab):
    with tab:
        st.subheader("Markdown Summaries")
        if os.path.isfile(MARKDOWN_CSV):
            try:
                md_df = pd.read_csv(MARKDOWN_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                all_md = "\n\n---\n\n".join(md_df["markdown_summary"].dropna().tolist())
                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "⬇️ Download All as .md",
                        all_md.encode("utf-8"),
                        file_name="incentive_summaries.md",
                        mime="text/markdown",
                        use_container_width=True,
                    )
                with c2:
                    with open(MARKDOWN_CSV, "rb") as f:
                        st.download_button(
                            "⬇️ Download Summaries CSV",
                            f,
                            file_name="markdown_summaries.csv",
                            use_container_width=True,
                        )
                for _, row in md_df.iterrows():
                    st.markdown(row["markdown_summary"])
                    st.markdown("---")
            except Exception as e:
                st.error(f"Could not load summaries: {e}")
        else:
            st.info("No summaries generated yet.")


def _render_errors_tab(tab):
    with tab:
        st.subheader("Error Log")
        if os.path.isfile(ERRORS_CSV):
            try:
                err_df = pd.read_csv(ERRORS_CSV, quoting=csv.QUOTE_ALL, on_bad_lines="skip")
                st.dataframe(err_df, use_container_width=True)
                with open(ERRORS_CSV, "rb") as f:
                    st.download_button("⬇️ Download Error Log", f, file_name="errors.csv", use_container_width=True)
            except Exception as e:
                st.error(f"Could not load errors: {e}")
        else:
            st.info("No errors logged.")


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: Upload Excel
# ═══════════════════════════════════════════════════════════════════════════════
if mode == "Upload Excel":
    tab_progress, tab_markdown, tab_errors = st.tabs(["📊 Progress", "📝 Summaries", "⚠️ Errors"])

    url_col_choice    = None
    parent_col_choice = None

    if uploaded_file:
        try:
            df_raw = pd.read_excel(uploaded_file)
            with tab_progress:
                st.subheader("File Preview")
                st.dataframe(df_raw.head(5), use_container_width=True)
                st.caption(f"{len(df_raw)} rows · {len(df_raw.columns)} columns")

                col_options = df_raw.columns.tolist()
                auto_url    = _find_url_column(col_options)
                c1, c2 = st.columns(2)
                with c1:
                    url_col_choice = st.selectbox(
                        "URL column *",
                        col_options,
                        index=col_options.index(auto_url) if auto_url else 0,
                        help="Column that contains the URLs to scrape",
                    )
                with c2:
                    parent_raw = st.selectbox(
                        "Parent URL column (optional)",
                        ["— none —"] + col_options,
                        index=0,
                        help="Optional: links a sub-page back to its parent domain",
                    )
                    parent_col_choice = None if parent_raw == "— none —" else parent_raw
        except Exception as e:
            st.error(f"Could not read Excel: {e}")

    if run_button:
        if not uploaded_file:
            st.error("Please upload an Excel file first.")
        elif url_col_choice is None:
            st.error("Select a URL column before running.")
        else:
            st.session_state.cancelled          = False
            st.session_state.excel_complete     = False
            st.session_state.excel_output_file  = None

            try:
                df_run = pd.read_excel(uploaded_file)
                # Rename chosen columns to what run_pipeline expects
                rename = {url_col_choice: "URLs"}
                if parent_col_choice:
                    rename[parent_col_choice] = "parent_url"
                df_run = df_run.rename(columns=rename)

                buf = io.BytesIO()
                df_run.to_excel(buf, index=False)
                buf.seek(0)

                with tab_progress:
                    progress_bar = st.progress(0)
                    status_text  = st.empty()
                    stats_text   = st.empty()

                success_count = [0]
                fail_count    = [0]

                def excel_progress(current, total, url="", message=""):
                    pct = current / total if total > 0 else 0
                    progress_bar.progress(pct)
                    status_text.markdown(f"**{message}**")
                    stats_text.markdown(
                        f"✅ `{success_count[0]}` succeeded &nbsp;|&nbsp; "
                        f"❌ `{fail_count[0]}` failed &nbsp;|&nbsp; "
                        f"🔗 `{url}`"
                    )

                output_file = run_pipeline(
                    mode="Upload Excel",
                    uploaded_file=buf,
                    temperature=temperature,
                    truncation_length=int(truncation_length),
                    progress_callback=excel_progress,
                    cancel_flag=lambda: st.session_state.cancelled,
                    provider=provider,
                    model=model_name,
                )
                st.session_state.excel_output_file = output_file
                st.session_state.excel_complete    = True

            except Exception as e:
                st.error(f"Pipeline failed: {e}")

    # ── Persistent results ────────────────────────────────────────────────────
    if st.session_state.excel_complete and st.session_state.excel_output_file:
        with tab_progress:
            if st.session_state.cancelled:
                st.warning("Cancelled — partial results below.")
            else:
                st.success("✅ Extraction complete.")
            with open(st.session_state.excel_output_file, "rb") as f:
                st.download_button(
                    "⬇️ Download Results CSV", f,
                    file_name="incentives_output.csv",
                    use_container_width=True,
                )
        _render_markdown_tab(tab_markdown)
        _render_errors_tab(tab_errors)


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: Single URL
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "Single URL":
    tab_progress, tab_markdown, tab_errors = st.tabs(["📊 Progress", "📝 Summaries", "⚠️ Errors"])

    if run_button:
        if not single_url_in or not single_url_in.strip():
            st.error("Enter a URL before running.")
        else:
            st.session_state.cancelled          = False
            st.session_state.single_complete    = False
            st.session_state.single_output_file = None

            try:
                df_single = pd.DataFrame({"URLs": [single_url_in.strip()]})
                buf = io.BytesIO()
                df_single.to_excel(buf, index=False)
                buf.seek(0)

                with tab_progress:
                    progress_bar = st.progress(0)
                    status_text  = st.empty()

                def single_progress(current, total, url="", message=""):
                    progress_bar.progress(current / total if total > 0 else 0)
                    status_text.markdown(f"**{message}**")

                output_file = run_pipeline(
                    mode="Upload Excel",
                    uploaded_file=buf,
                    temperature=temperature,
                    truncation_length=int(truncation_length),
                    progress_callback=single_progress,
                    cancel_flag=lambda: st.session_state.cancelled,
                    provider=provider,
                    model=model_name,
                )
                st.session_state.single_output_file = output_file
                st.session_state.single_complete    = True

            except Exception as e:
                st.error(f"Pipeline failed: {e}")

    # ── Persistent results ────────────────────────────────────────────────────
    if st.session_state.single_complete and st.session_state.single_output_file:
        with tab_progress:
            st.success("✅ Extraction complete.")
            with open(st.session_state.single_output_file, "rb") as f:
                st.download_button(
                    "⬇️ Download Results CSV", f,
                    file_name="incentives_output.csv",
                    use_container_width=True,
                )
        _render_markdown_tab(tab_markdown)
        _render_errors_tab(tab_errors)


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: Upload Markdown
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "Upload Markdown":
    tab_progress, tab_markdown, tab_errors = st.tabs(["📊 Progress", "📝 Summaries", "⚠️ Errors"])

    if uploaded_md:
        with tab_progress:
            st.subheader("File Preview")
            preview_text = uploaded_md.read().decode("utf-8", errors="replace")
            uploaded_md.seek(0)
            st.text_area("First 3000 characters", preview_text[:3000], height=200, disabled=True)

    if run_button:
        if not uploaded_md:
            st.error("Upload a markdown or text file first.")
        else:
            st.session_state.cancelled       = False
            st.session_state.md_complete     = False
            st.session_state.md_output_file  = None
            st.session_state.md_input_text   = None
            st.session_state.md_input_name   = None

            try:
                from modules.processor import process_text
                from modules.exporter import export_to_csv, append_markdown_entry
                import datetime as dt

                markdown_content = uploaded_md.read().decode("utf-8", errors="replace")

                with tab_progress:
                    st.info("Running LLM extraction on uploaded file…")

                structured = process_text(
                    markdown_content,
                    source_name=uploaded_md.name,
                    temperature=temperature,
                    provider=provider,
                    model_name=model_name,
                )
                structured["source_url"]           = uploaded_md.name
                structured["extraction_timestamp"] = dt.datetime.utcnow().isoformat()
                append_markdown_entry(structured, MARKDOWN_CSV)
                output_file = export_to_csv([structured])

                st.session_state.md_output_file = output_file
                st.session_state.md_input_text  = markdown_content
                st.session_state.md_input_name  = uploaded_md.name
                st.session_state.md_complete    = True

            except Exception as e:
                st.error(f"Extraction failed: {e}")

    # ── Persistent results ────────────────────────────────────────────────────
    if st.session_state.md_complete and st.session_state.md_output_file:
        with tab_progress:
            st.success("✅ Extraction complete.")
            c1, c2 = st.columns(2)
            with c1:
                with open(st.session_state.md_output_file, "rb") as f:
                    st.download_button(
                        "⬇️ Download Results CSV", f,
                        file_name="incentives_output.csv",
                        use_container_width=True,
                    )
            with c2:
                if st.session_state.md_input_text:
                    st.download_button(
                        "⬇️ Download Original Markdown", 
                        st.session_state.md_input_text.encode("utf-8"),
                        file_name=st.session_state.md_input_name or "input.md",
                        mime="text/markdown",
                        use_container_width=True,
                    )
        _render_markdown_tab(tab_markdown)
        _render_errors_tab(tab_errors)


# ═══════════════════════════════════════════════════════════════════════════════
# MODE: City URL Discovery
# ═══════════════════════════════════════════════════════════════════════════════
elif mode == "City URL Discovery":
    st.subheader("Discover New Utility URLs by State")
    st.caption(
        "Searches OpenSERP for electric utility and cooperative websites by state. "
        "Requires OpenSERP running locally — start it with: "
        "`cd modules/searching && .\\openserp.exe serve`"
    )

    # ── OpenSERP health check ─────────────────────────────────────────────────
    with st.expander("🔌 OpenSERP connection check", expanded=False):
        check_url = st.text_input("OpenSERP URL to test", value="http://localhost:7070", key="check_url")
        if st.button("Check connection"):
            import requests as _req
            try:
                r = _req.get(f"{check_url}/google/search", params={"text": "test", "limit": 1}, timeout=5)
                if r.status_code == 200:
                    st.success("✅ OpenSERP is reachable.")
                else:
                    st.warning(f"Responded with status {r.status_code} — may not be working correctly.")
            except Exception as ex:
                st.error(
                    f"Could not reach OpenSERP: {ex}\n\n"
                    f"Start it with:\n```\ncd modules/searching\n.\\openserp.exe serve\n```"
                )

    st.markdown("---")

    disc_col1, disc_col2 = st.columns(2)
    with disc_col1:
        selected_states = st.multiselect("States to search", VALID_STATES, default=["Texas"])
        openserp_url    = st.text_input("OpenSERP URL", value="http://localhost:7070")
        engine          = st.selectbox("Search engine", ["google", "bing", "duckduckgo"], index=0)
    with disc_col2:
        num_results = st.slider("Results per query", min_value=3, max_value=15, value=8)
        db_file     = st.file_uploader(
            "Existing URL database for deduplication (optional)",
            type=["xlsx"],
            key="db_upload",
        )

    with st.expander("Search topics"):
        topics_text = st.text_area(
            "One topic per line",
            value="\n".join(DISCOVERY_TOPICS),
            height=300,
        )

    if st.button("▶ Run Discovery", disabled=not selected_states, type="primary"):
        # Clear previous discovery results when a new run starts
        st.session_state.discovery_result = None

        progress_bar = st.progress(0)
        status_text  = st.empty()

        def discovery_progress(current, total, url="", message=""):
            progress_bar.progress(current / total if total > 0 else 0)
            status_text.markdown(f"**{message}**")

        discovered = get_urls_from_discovery(
            states=selected_states,
            openserp_url=openserp_url,
            engine=engine,
            num_results=num_results,
            existing_db=db_file,
            progress_callback=discovery_progress,
        )
        st.session_state.discovery_result = discovered

    # ── Persistent discovery results ──────────────────────────────────────────
    discovered = st.session_state.discovery_result
    if discovered is not None:
        st.success(f"Found **{len(discovered)}** new utility URLs.")

        if discovered:
            st.subheader("Discovered URLs")
            df_disc = pd.DataFrame([{
                "State":         r["state"],
                "URL":           r["url"],
                "Page Title":    r.get("title", ""),
                "Discovered At": r.get("discovered_at", ""),
            } for r in discovered])
            st.dataframe(df_disc, use_container_width=True)

            # ── Excel download ────────────────────────────────────────────────
            wb  = Workbook()
            ws  = wb.active
            ws.title = "Discovered URLs"
            for col, w in {"A": 16, "B": 30, "C": 60, "D": 45, "E": 70, "F": 22}.items():
                ws.column_dimensions[col].width = w
            ws.freeze_panes = "A2"

            for c, h in enumerate(["State", "Search Query", "URL", "Page Title", "Description", "Discovered At"], 1):
                cell           = ws.cell(row=1, column=c, value=h)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                cell.fill      = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin()

            for r_idx, row in enumerate(discovered, 2):
                for c, val in enumerate([
                    row.get("state", ""), row.get("query", ""), row.get("url", ""),
                    row.get("title", ""), row.get("description", ""), row.get("discovered_at", ""),
                ], 1):
                    cell           = ws.cell(row=r_idx, column=c, value=val)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = _thin()
                    cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 4, 5)))
                    if c == 1:
                        cell.fill = PatternFill("solid", start_color="D6E4F0")
                        cell.font = Font(name="Arial", bold=True, size=10, color="1F4E79")

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "⬇️ Download Discovered URLs Excel",
                    buf.read(),
                    file_name=f"utility_urls_discovered_{datetime.date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with c2:
                url_list = "\n".join(r["url"] for r in discovered)
                st.download_button(
                    "⬇️ Download as plain URL list (.txt)",
                    url_list.encode("utf-8"),
                    file_name="discovered_urls.txt",
                    mime="text/plain",
                    use_container_width=True,
                )

    # ── Merge sub-section ─────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Merge Discovered URLs into Existing Database")
    st.caption("Domain-level deduplication. Merges a discovered file into your existing URL database.")

    merge_col1, merge_col2 = st.columns(2)
    with merge_col1:
        merge_db   = st.file_uploader("Existing URL database",  type=["xlsx"], key="merge_db")
    with merge_col2:
        merge_disc = st.file_uploader("Discovered URLs file",   type=["xlsx"], key="merge_disc")

    if merge_db and merge_disc:
        from openpyxl import load_workbook as lw

        def _load_urls(f) -> list:
            wb = lw(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            urls = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and cell.startswith("http"):
                        urls.append(cell.strip())
            wb.close()
            return urls

        def _load_disc_rows(f) -> list:
            wb      = lw(io.BytesIO(f.read()), read_only=True)
            ws      = wb.active
            rows    = []
            headers = None
            for row in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(headers, row))
                url    = str(record.get("URL", "") or "").strip()
                if url.startswith("http"):
                    rows.append(record)
            wb.close()
            return rows

        existing_urls    = _load_urls(merge_db)
        existing_domains = set(_extract_domain(u) for u in existing_urls if _extract_domain(u))
        discovered_rows  = _load_disc_rows(merge_disc)

        seen     = set(existing_domains)
        new_rows = []
        skipped  = 0
        for row in discovered_rows:
            url    = str(row.get("URL", "")).strip()
            domain = _extract_domain(url)
            if not domain:
                continue
            if domain in seen:
                skipped += 1
            else:
                seen.add(domain)
                new_rows.append(row)

        st.write(f"**Existing:** {len(existing_urls)} URLs &nbsp;|&nbsp; **Discovered:** {len(discovered_rows)} rows")
        st.write(f"**New after dedup:** {len(new_rows)} &nbsp;|&nbsp; **Skipped (duplicate domain):** {skipped}")

        if new_rows and st.button("▶ Merge & Download", type="primary"):
            wb        = _build_merged_workbook(existing_urls, new_rows)
            buf       = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.success(f"Merged: {len(existing_urls)} existing + {len(new_rows)} new = {len(existing_urls) + len(new_rows)} total")
            st.session_state.merge_result = True
            st.download_button(
                "⬇️ Download Merged Database",
                buf.read(),
                file_name=f"Relevant_URLs_merged_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
